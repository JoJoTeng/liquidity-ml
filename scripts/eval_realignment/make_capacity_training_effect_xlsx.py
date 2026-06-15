"""Summarize the section 4.2 capacity-portfolio training effect into one workbook.

Reads the metrics CSVs written by 42_signal_weighted_capacity_portfolio.py and
writes a single plain .xlsx with one sheet per model. Within each model sheet the
results are stacked as three labelled panels covering every fitted weight spec:

  * "Gross SR"   - gross annualized Sharpe levels (standard, weighted) and their
                   difference (AUM-independent).
  * "Net SR"     - net annualized Sharpe levels (standard, weighted) and their
                   difference (weighted - standard) per AUM scenario; a trailing
                   '*' on the difference marks p < 0.05 from the one-sided
                   Ledoit-Wolf bootstrap Sharpe test (weighted > standard).
  * "dCE (g=5)"  - annualized net certainty-equivalent training effect at
                   gamma = 5 per AUM scenario.

Effect sizes and gross levels only; no inference. Reads outputs, writes a
workbook; nothing in the analysis cache is modified.

    python scripts/eval_realignment/make_capacity_training_effect_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.workbook import Workbook

REPO = Path(__file__).resolve().parent.parent.parent
ANALYSIS = REPO / "outputs" / "eval_realignment" / "analysis"
OUT = REPO / "outputs" / "eval_realignment" / "tables" / "capacity_portfolio_training_effect.xlsx"

# (model directory, sheet label)
MODELS = [("elastic_net", "ElasticNet"), ("xgboost", "XGBoost"), ("neural_network", "NeuralNet")]

# (spec directory, display label) in the formal fitted order.
SPECS = [
    ("dolvol", "dolvol"),
    ("softmax_rank_lam2", "softmax λ2"),
    ("softmax_rank_lam3", "softmax λ3"),
    ("tc_10m", "TC $10M"),
    ("tc_100m", "TC $100M"),
    ("tc_500m", "TC $500M"),
    ("tc_1000m", "TC $1B"),
    ("tc_rank_lam3_10m", "TCrank $10M"),
    ("tc_rank_lam3_100m", "TCrank $100M"),
    ("tc_rank_lam3_500m", "TCrank $500M"),
    ("tc_rank_lam3_1000m", "TCrank $1B"),
]

# (AUM file token, display label); PropTC (no market impact) shown first.
AUMS = [("PropTC", "PropTC"), ("100M", "100M"), ("500M", "500M"), ("1B", "1B")]


def _metrics(model_dir: str, spec_dir: str, aum: str) -> pd.DataFrame | None:
    f = ANALYSIS / model_dir / spec_dir / f"capacity_portfolio_metrics_{aum}.csv"
    if not f.exists():
        return None
    return pd.read_csv(f).set_index("row_type")


def _first_metrics(model_dir: str, spec_dir: str) -> pd.DataFrame | None:
    for aum, _ in AUMS:
        d = _metrics(model_dir, spec_dir, aum)
        if d is not None:
            return d
    return None


def _r(x) -> float | None:
    return None if x is None or pd.isna(x) else round(float(x), 4)


def _gross_panel(model_dir: str) -> pd.DataFrame:
    rows = []
    for spec_dir, slab in SPECS:
        d = _first_metrics(model_dir, spec_dir)
        if d is None or not {"weighted", "standard"} <= set(d.index):
            rows.append({"Spec": slab, "SR std": None, "SR wt": None, "ΔSR": None})
            continue
        gs = d.loc["standard", "gross_sr_annual"]
        gw = d.loc["weighted", "gross_sr_annual"]
        rows.append({"Spec": slab, "SR std": _r(gs), "SR wt": _r(gw), "ΔSR": _r(gw - gs)})
    return pd.DataFrame(rows)[["Spec", "SR std", "SR wt", "ΔSR"]]


def _net_panel(model_dir: str) -> pd.DataFrame:
    """Net Sharpe levels (standard, weighted) and their difference, per AUM.

    A trailing '*' on the difference marks p < 0.05 from the one-sided
    Ledoit-Wolf bootstrap Sharpe test (weighted > standard) on net returns.
    """
    cols = ["Spec"]
    for _, alab in AUMS:
        cols += [f"{alab} std", f"{alab} wt", f"{alab} Δ"]
    rows = []
    for spec_dir, slab in SPECS:
        row = {"Spec": slab}
        for atok, alab in AUMS:
            d = _metrics(model_dir, spec_dir, atok)
            ns = nw = delta = None
            star = ""
            if d is not None and {"weighted", "standard"} <= set(d.index):
                ns = d.loc["standard", "net_sr_annual"]
                nw = d.loc["weighted", "net_sr_annual"]
                delta = nw - ns
                if "difference" in d.index:
                    p = d.loc["difference", "net_sr_diff_pval"]
                    if pd.notna(p) and float(p) < 0.05:
                        star = "*"
            row[f"{alab} std"] = _r(ns)
            row[f"{alab} wt"] = _r(nw)
            dv = _r(delta)
            row[f"{alab} Δ"] = None if dv is None else (f"{dv}{star}" if star else dv)
        rows.append(row)
    return pd.DataFrame(rows)[cols]


def _ce_panel(model_dir: str) -> pd.DataFrame:
    """Net certainty-equivalent training effect (gamma = 5), per AUM."""
    cols = ["Spec"] + [alab for _, alab in AUMS]
    rows = []
    for spec_dir, slab in SPECS:
        row = {"Spec": slab}
        for atok, alab in AUMS:
            d = _metrics(model_dir, spec_dir, atok)
            v = None
            if d is not None and "difference" in d.index:
                v = d.loc["difference", "net_ce_diff_annual_g5"]
            row[alab] = _r(v)
        rows.append(row)
    return pd.DataFrame(rows)[cols]


def _dump(ws, start_row: int, title: str, df: pd.DataFrame) -> int:
    """Write a titled panel starting at 1-based ``start_row``; return next free row."""
    ws.cell(row=start_row, column=1, value=title)
    header_row = start_row + 1
    for c, name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=c, value=name)
    for i, (_, rec) in enumerate(df.iterrows()):
        for c, name in enumerate(df.columns, start=1):
            ws.cell(row=header_row + 1 + i, column=c, value=rec[name])
    return header_row + 1 + len(df) + 1  # one blank row before the next panel


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for model_dir, sheet in MODELS:
        ws = wb.create_sheet(title=sheet)
        r = 1
        r = _dump(ws, r, "Gross SR", _gross_panel(model_dir))
        r = _dump(ws, r, "Net SR  (Δ: * = p < 0.05, one-sided bootstrap)", _net_panel(model_dir))
        r = _dump(ws, r, "ΔCE (γ=5)", _ce_panel(model_dir))
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(REPO)} ({len(MODELS)} model sheets, 3 panels each)")


if __name__ == "__main__":
    main()
