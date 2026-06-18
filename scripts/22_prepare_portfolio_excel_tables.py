"""
Prepare paper-style Excel tables from formal portfolio outputs.

This script is intentionally lightweight: it does not recompute portfolios. It
reads the CSV outputs produced by ``21e_formal_portfolio_decomposition.py`` and
writes formatted Excel tables for reporting. Multi-table exports use one
workbook per model, one worksheet per weighted-training spec, and one repeated
subtable per AUM case.

Tables 13 and 14 are prediction-quantile based. Use ``--stock-universe nyse``
to read NYSE-only 21e portfolio outputs.
"""

from __future__ import annotations

import argparse
import os
import sys
import tempfile
from pathlib import Path
from typing import Sequence

_runtime_cache = Path(tempfile.gettempdir()) / "liquidity_ml_cache"
_runtime_cache.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_runtime_cache / "matplotlib"))
os.environ.setdefault("XDG_CACHE_HOME", str(_runtime_cache / "xdg"))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.analysis.formal.common import is_proportional_tc_scenario  # noqa: E402
from src.analysis.formal.script_utils import (  # noqa: E402
    STOCK_UNIVERSE_CHOICES,
    resolve_stock_universes,
    stock_universe_read_dir,
)
from src.config import load_config  # noqa: E402
from src.evaluation.statistics import sharpe_ratio  # noqa: E402


MODEL_CHOICES = ["elastic_net", "xgboost", "neural_network"]
TRAINING_AUMS = [10, 100, 500, 1000]
DEFAULT_AUMS = ["proportional_tc", 100, 500, 1000]
PRIMARY_WEIGHT_SPECS = [
    "dolvol",
    "softmax_rank_lam2",
    "tc_500m",
    "tc_rank_lam3_500m",
]
WEIGHT_SPEC_CHOICES = [
    "dolvol",
    "softmax_rank_lam2",
    "softmax_rank_lam3",
    "tc_10m",
    "tc_100m",
    "tc_500m",
    "tc_1000m",
] + [
    f"tc_rank_lam{lam}_{aum}m"
    for lam in (3,)
    for aum in TRAINING_AUMS
]
PORTFOLIO_RUN_CHOICES = [
    "prediction_quantile",
    "prediction_quantile_value_weight",
    "prediction_quantile_dolvol_weight",
    "prediction_quantile_dolvol_weight_liq60",
]


def parse_args():
    parser = argparse.ArgumentParser(
        description="Prepare formatted Excel tables from formal portfolio outputs"
    )
    parser.add_argument(
        "--table",
        default="table13",
        choices=["table13", "table14"],
        help=(
            "Report table to export. table13 is the standalone Q1-Q5 "
            "prediction-quantile portfolio table; table14 is the 2x3 "
            "decomposition for each standalone prediction quantile. "
            "(The Q5-Q1 long-short decomposition lives in "
            "scripts/22b_table12_two_sided.py.)"
        ),
    )
    parser.add_argument(
        "--model",
        default="elastic_net",
        help=(
            "Model family to export. Use one model, a comma-separated list, "
            "or 'all'."
        ),
    )
    parser.add_argument(
        "--weight-spec",
        default="primary",
        help=(
            "Formal weight-spec folder to export. Use one spec, a "
            "comma-separated list, 'primary' for the four main specs, or 'all'."
        ),
    )
    parser.add_argument(
        "--portfolio-run",
        default="prediction_quantile_dolvol_weight_liq60",
        choices=[*PORTFOLIO_RUN_CHOICES, "both"],
        help="Portfolio output folder produced by 21e, or 'both'.",
    )
    parser.add_argument(
        "--stock-universe",
        default="full_sample",
        choices=STOCK_UNIVERSE_CHOICES,
        help=(
            "Stock universe used by 21e portfolio construction. 'full_sample' "
            "reads legacy full-sample outputs when no namespaced folder exists; "
            "'nyse' reads stock_universe/nyse; 'both' writes separate table "
            "workbooks for full_sample and nyse."
        ),
    )
    parser.add_argument(
        "--aum",
        default="500",
        help=(
            "AUM in $M for the net Sharpe column, or proportional_tc for "
            "spread/2 only. Use one value, a comma-separated list, or 'all'. "
            "Default is 500."
        ),
    )
    parser.add_argument(
        "--skip-missing",
        action="store_true",
        help=(
            "Skip missing 21e output combinations instead of failing. Useful "
            "when exporting all models before every model has been run."
        ),
    )
    parser.add_argument(
        "--output",
        default=None,
        help=(
            "Optional output .xlsx path or root directory. Default writes under "
            "outputs/formalanalysis/tables/{model}/{stock_universe}/."
        ),
    )
    return parser.parse_args()


def _aum_label(aum_millions: int | str) -> str:
    if is_proportional_tc_scenario(aum_millions):
        return "PropTC"
    if aum_millions < 1000:
        return f"{aum_millions}M"
    if aum_millions % 1000 == 0:
        return f"{aum_millions // 1000}B"
    return f"{aum_millions / 1000:.1f}B"


def _parse_selection(value: str, choices: Sequence[str], name: str) -> list[str]:
    if value.lower() == "all":
        return list(choices)

    selected = [item.strip() for item in value.split(",") if item.strip()]
    invalid = sorted(set(selected).difference(choices))
    if invalid:
        valid = ", ".join(choices)
        raise ValueError(f"Unknown {name}: {invalid}. Valid values are: {valid}")
    if not selected:
        raise ValueError(f"No {name} selected.")
    return selected


def _parse_weight_specs(value: str) -> list[str]:
    if value.lower() == "primary":
        return list(PRIMARY_WEIGHT_SPECS)
    return _parse_selection(value, WEIGHT_SPEC_CHOICES, "weight spec")


def _parse_portfolio_runs(value: str) -> list[str]:
    if value.lower() == "both":
        return list(PORTFOLIO_RUN_CHOICES)
    return _parse_selection(value, PORTFOLIO_RUN_CHOICES, "portfolio run")


def _parse_aums(value: str) -> list[int | str]:
    if value.lower() == "all":
        return list(DEFAULT_AUMS)

    out: list[int | str] = []
    for raw in value.split(","):
        raw = raw.strip()
        if not raw:
            continue
        if is_proportional_tc_scenario(raw):
            out.append("proportional_tc")
            continue
        try:
            out.append(int(raw))
        except ValueError as exc:
            raise ValueError(
                f"AUM must be an integer in $M or proportional_tc, got {raw!r}"
            ) from exc
    if not out:
        raise ValueError("No AUM selected.")
    return out


def _spec_title(weight_spec: str) -> str:
    if weight_spec.startswith("tc_rank_lam"):
        token = weight_spec.removeprefix("tc_rank_lam")
        try:
            lam_token, aum_token = token.rsplit("_", 1)
            lam = float(lam_token.replace("m", "-").replace("p", "."))
            aum_m = int(aum_token.removesuffix("m"))
            return f"TC-Rank Weights, lambda={lam:g}, training AUM ${_aum_label(aum_m)}"
        except ValueError:
            return weight_spec

    labels = {
        "dolvol": "Dollar-Volume Weights",
        "softmax_rank_lam2": "Softmax-Rank Weights, lambda=2",
        "softmax_rank_lam3": "Softmax-Rank Weights, lambda=3",
        "tc_10m": "TC Weights, $10M",
        "tc_100m": "TC Weights, $100M",
        "tc_500m": "TC Weights, $500M",
        "tc_1000m": "TC Weights, $1B",
    }
    return labels.get(weight_spec, weight_spec)


def _model_title(model: str) -> str:
    labels = {
        "elastic_net": "ElasticNet",
        "xgboost": "XGBoost",
        "neural_network": "Neural Network",
    }
    return labels.get(model, model)


def _portfolio_run_title(portfolio_run: str) -> str:
    labels = {
        "prediction_quantile": "Equal-weighted prediction quantiles",
        "prediction_quantile_value_weight": "Value-weighted prediction quantiles",
        "prediction_quantile_dolvol_weight": "Dollar-volume-weighted prediction quantiles",
        "prediction_quantile_dolvol_weight_liq60": (
            "Dollar-volume-weighted prediction quantiles, top-60% liquid universe"
        ),
    }
    return labels.get(portfolio_run, portfolio_run.replace("_", "-"))


def _stock_universe_title(stock_universe: str) -> str:
    labels = {
        "full_sample": "Full stock universe",
        "nyse": "NYSE stock universe",
    }
    return labels.get(stock_universe, stock_universe.replace("_", " ").title())


def _portfolio_output_dir(
    analysis_dir: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    stock_universe: str,
) -> Path:
    """Return the 21e output directory for a portfolio run and stock universe."""
    base_dir = analysis_dir / model / weight_spec / portfolio_run
    return stock_universe_read_dir(base_dir, stock_universe)


def _selection_label(
    selected: Sequence[object],
    choices: Sequence[object],
    all_label: str,
    multi_label: str,
) -> str:
    if list(selected) == list(choices):
        return all_label
    if len(selected) == 1:
        value = selected[0]
        if isinstance(value, int) or is_proportional_tc_scenario(value):
            return _aum_label(value)
        return str(value)
    return f"{len(selected)}_{multi_label}"


def _aum_title(aum_millions: int | str) -> str:
    """Return a report title fragment for an AUM or proportional-cost case."""
    if is_proportional_tc_scenario(aum_millions):
        return "proportional TC only"
    return f"${_aum_label(aum_millions)} AUM"


def _set_table13_widths(ws) -> None:
    widths = {
        "A": 8,
        "B": 24,
        "C": 20,
        "D": 10,
        "E": 13,
        "F": 13,
        "G": 10,
        "H": 11,
        "I": 11,
        "J": 11,
        "K": 10,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _set_table12_widths(ws) -> None:
    widths = {
        "A": 34,
        "B": 20,
        "C": 20,
        "D": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _write_table12_block(
    ws,
    table: dict[str, float],
    start_row: int,
    title: str,
    subtitle: str | None = None,
) -> int:
    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="404040")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=4)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = title_row + 1
    if subtitle:
        subtitle_row = title_row + 1
        ws.merge_cells(
            start_row=subtitle_row,
            start_column=1,
            end_row=subtitle_row,
            end_column=4,
        )
        subtitle_cell = ws.cell(row=subtitle_row, column=1, value=subtitle)
        subtitle_cell.font = Font(italic=True, size=10, color="555555")
        subtitle_cell.alignment = Alignment(horizontal="center")
        header_row = title_row + 2

    headers = ["", "Sort on r_hat", "TC-aware sort", "TC-target score"]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=medium)
        cell.fill = header_fill

    panel_a_row = header_row + 1
    ws.merge_cells(
        start_row=panel_a_row,
        start_column=1,
        end_row=panel_a_row,
        end_column=4,
    )
    cell = ws.cell(row=panel_a_row, column=1, value="Panel A: Net Sharpe Ratios")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=thin)

    net_rows = [
        (
            "Standard training",
            ("1A", table["SR_net_1A"]),
            ("1B", table["SR_net_1B"]),
            ("1C", table["SR_net_1C"]),
        ),
        (
            "Weighted training",
            ("2A", table["SR_net_2A"]),
            ("2B", table["SR_net_2B"]),
            ("2C", table["SR_net_2C"]),
        ),
    ]
    row_idx = panel_a_row + 1
    for label, col_b, col_c, col_d in net_rows:
        ws.cell(row=row_idx, column=1, value=label)
        for col_idx, (cell_name, value) in enumerate([col_b, col_c, col_d], start=2):
            ws.cell(row=row_idx, column=col_idx, value=f"{cell_name}: {value:.3f}")
        for col_idx in range(1, 5):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center"
            )
        row_idx += 1

    panel_b_row = row_idx
    ws.merge_cells(
        start_row=panel_b_row,
        start_column=1,
        end_row=panel_b_row,
        end_column=4,
    )
    cell = ws.cell(row=panel_b_row, column=1, value="Panel B: 2x2 Decomposition")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=medium)

    decomp_rows = [
        ("Training effect (2A - 1A)", table["training_effect"], "0.000"),
        ("Portfolio effect (1B - 1A)", table["portfolio_effect"], "0.000"),
        ("Total effect (2B - 1A)", table["total_effect"], "0.000"),
        ("Interaction", table["interaction"], "0.000"),
        ("Training share (%)", table["training_share_pct"], "0.0"),
    ]
    row_idx = panel_b_row + 1
    for label, value, fmt in decomp_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.number_format = fmt
        value_cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    panel_c_row = row_idx
    ws.merge_cells(
        start_row=panel_c_row,
        start_column=1,
        end_row=panel_c_row,
        end_column=4,
    )
    cell = ws.cell(row=panel_c_row, column=1, value="Panel C: TC-Target Score Effects")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=medium)

    target_rows = [
        ("Standard: 1C - 1A", table["tc_target_standard_vs_prediction"]),
        ("Standard: 1C - 1B", table["tc_target_standard_vs_tc_sort"]),
        ("Weighted: 2C - 2A", table["tc_target_weighted_vs_prediction"]),
        ("Weighted: 2C - 2B", table["tc_target_weighted_vs_tc_sort"]),
        ("TC-target training effect (2C - 1C)", table["tc_target_training_effect"]),
        ("TC-target total effect (2C - 1A)", table["tc_target_total_effect"]),
    ]
    row_idx = panel_c_row + 1
    for label, value in target_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.number_format = "0.000"
        value_cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    bottom_row = row_idx
    for col_idx in range(1, 5):
        ws.cell(row=bottom_row, column=col_idx).border = Border(top=medium)

    return bottom_row + 1


def _safe_spec_sheet_title(weight_spec: str) -> str:
    return weight_spec[:31]


def load_prediction_quantile_table(
    analysis_dir: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    aum_millions: int | str,
    stock_universe: str = "full_sample",
) -> pd.DataFrame:
    """Load 21e Q1-Q5 prediction-quantile output and make Table 13 rows."""
    spec_dir = _portfolio_output_dir(
        analysis_dir,
        model,
        weight_spec,
        portfolio_run,
        stock_universe,
    )
    path = spec_dir / f"prediction_quantile_timeseries_{_aum_label(aum_millions)}.csv"
    if not path.exists():
        raise FileNotFoundError(
            f"Missing 21e prediction-quantile output:\n{path}\n"
            "Run 21e before preparing Table 13."
        )

    raw = pd.read_csv(path)
    required_cols = {
        "yyyymm",
        "cell",
        "training_model",
        "portfolio_sort",
        "prediction_quantile",
        "gross_return",
        "transaction_cost",
        "net_return",
        "turnover",
        "n_stocks",
    }
    missing_cols = required_cols.difference(raw.columns)
    if missing_cols:
        raise ValueError(f"{path} missing columns: {sorted(missing_cols)}")

    rows = []
    cell_order = {"1A": 0, "1B": 1, "1C": 2, "2A": 3, "2B": 4, "2C": 5}
    for (cell, q), group in raw.groupby(["cell", "prediction_quantile"], sort=False):
        group = group.sort_values("yyyymm")
        turnover = group["turnover"].iloc[1:].mean() if len(group) > 1 else float("nan")
        rows.append(
            {
                "cell": cell,
                "training_model": _training_model_label(
                    str(group["training_model"].iloc[0])
                ),
                "portfolio_sort": _portfolio_sort_label(
                    str(group["portfolio_sort"].iloc[0])
                ),
                "prediction_quantile": f"Q{int(q)}",
                "gross_return_pct": group["gross_return"].mean() * 100.0,
                "net_return_pct": group["net_return"].mean() * 100.0,
                "transaction_cost_pct": group["transaction_cost"].mean() * 100.0,
                "gross_sr": sharpe_ratio(group["gross_return"], annualize=True),
                "net_sr": sharpe_ratio(group["net_return"], annualize=True),
                "turnover": turnover,
                "avg_n_stocks": group["n_stocks"].mean(),
                "_cell_order": cell_order.get(cell, 99),
                "_quantile_order": int(q),
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return out
    out = out.sort_values(["_cell_order", "_quantile_order"]).drop(
        columns=["_cell_order", "_quantile_order"]
    )
    return out


def _load_prediction_quantile_timeseries(
    analysis_dir: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    aum_millions: int | str,
    stock_universe: str = "full_sample",
) -> pd.DataFrame:
    """Load the 21e monthly Q1-Q5 prediction-quantile time series."""
    spec_dir = _portfolio_output_dir(
        analysis_dir,
        model,
        weight_spec,
        portfolio_run,
        stock_universe,
    )
    path = spec_dir / f"prediction_quantile_timeseries_{_aum_label(aum_millions)}.csv"
    if not path.exists():
        raise FileNotFoundError(
            f"Missing 21e prediction-quantile output:\n{path}\n"
            "Run 21e before preparing prediction-quantile tables."
        )

    raw = pd.read_csv(path)
    required_cols = {
        "yyyymm",
        "cell",
        "prediction_quantile",
        "gross_return",
        "transaction_cost",
        "net_return",
    }
    missing_cols = required_cols.difference(raw.columns)
    if missing_cols:
        raise ValueError(f"{path} missing columns: {sorted(missing_cols)}")
    return raw


def _quantile_decomposition_from_timeseries(raw: pd.DataFrame) -> dict[int, dict[str, float]]:
    """Build Table-12-style 2x3 summaries for each standalone prediction quantile."""
    out: dict[int, dict[str, float]] = {}
    cells = ["1A", "1B", "1C", "2A", "2B", "2C"]

    for quantile in sorted(raw["prediction_quantile"].dropna().astype(int).unique()):
        q_raw = raw[raw["prediction_quantile"].astype(int) == quantile]
        cell_frames: dict[str, pd.DataFrame] = {}
        common_months: set[int] | None = None

        for cell in cells:
            frame = q_raw[q_raw["cell"] == cell].sort_values("yyyymm")
            if frame.empty:
                continue
            cell_frames[cell] = frame
            months = set(frame["yyyymm"].astype(int))
            common_months = months if common_months is None else common_months & months

        if set(cell_frames) != set(cells) or not common_months:
            continue

        aligned = {
            cell: frame[frame["yyyymm"].astype(int).isin(common_months)].sort_values("yyyymm")
            for cell, frame in cell_frames.items()
        }

        net_sr_annual = {
            cell: sharpe_ratio(aligned[cell]["net_return"], annualize=True)
            for cell in cells
        }

        training_effect = net_sr_annual["2A"] - net_sr_annual["1A"]
        portfolio_effect = net_sr_annual["1B"] - net_sr_annual["1A"]
        total_effect = net_sr_annual["2B"] - net_sr_annual["1A"]
        interaction = total_effect - training_effect - portfolio_effect
        training_share_pct = (
            training_effect / total_effect * 100.0
            if abs(total_effect) > 1e-10
            else float("nan")
        )

        out[int(quantile)] = {
            "SR_net_1A": net_sr_annual["1A"],
            "SR_net_1B": net_sr_annual["1B"],
            "SR_net_1C": net_sr_annual["1C"],
            "SR_net_2A": net_sr_annual["2A"],
            "SR_net_2B": net_sr_annual["2B"],
            "SR_net_2C": net_sr_annual["2C"],
            "training_effect": training_effect,
            "portfolio_effect": portfolio_effect,
            "total_effect": total_effect,
            "interaction": interaction,
            "training_share_pct": training_share_pct,
            "tc_target_standard_vs_prediction": net_sr_annual["1C"] - net_sr_annual["1A"],
            "tc_target_standard_vs_tc_sort": net_sr_annual["1C"] - net_sr_annual["1B"],
            "tc_target_weighted_vs_prediction": net_sr_annual["2C"] - net_sr_annual["2A"],
            "tc_target_weighted_vs_tc_sort": net_sr_annual["2C"] - net_sr_annual["2B"],
            "tc_target_training_effect": net_sr_annual["2C"] - net_sr_annual["1C"],
            "tc_target_total_effect": net_sr_annual["2C"] - net_sr_annual["1A"],
        }

    return out


def load_prediction_quantile_decomposition_tables(
    analysis_dir: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    aum_millions: int | str,
    stock_universe: str = "full_sample",
) -> dict[int, dict[str, float]]:
    """Load 21e Q1-Q5 time series and make Table-14 quantile decomposition rows."""
    raw = _load_prediction_quantile_timeseries(
        analysis_dir=analysis_dir,
        model=model,
        weight_spec=weight_spec,
        portfolio_run=portfolio_run,
        aum_millions=aum_millions,
        stock_universe=stock_universe,
    )
    tables = _quantile_decomposition_from_timeseries(raw)
    if not tables:
        raise ValueError(
            f"No complete Q1-Q5 2x3 quantile decompositions for "
            f"{model}/{weight_spec}/{portfolio_run}/{stock_universe}/"
            f"{_aum_label(aum_millions)}"
        )
    return tables


def _training_model_label(value: str) -> str:
    labels = {
        "standard": "Standard",
        "weighted": "Weighted",
        "standard_tc_target": "Standard TC-target",
        "weighted_tc_target": "Weighted TC-target",
    }
    return labels.get(value, value.replace("_", " ").title())


def _portfolio_sort_label(value: str) -> str:
    labels = {
        "prediction": "Sort on r_hat",
        "tc_net_score": "TC-aware sort",
        "tc_target_score": "TC-target score",
    }
    return labels.get(value, value.replace("_", " ").title())


def _load_model_spec_quantile_tables(
    analysis_dir: Path,
    model: str,
    weight_specs: Sequence[str],
    portfolio_run: str,
    aums: Sequence[int | str],
    skip_missing: bool,
    stock_universe: str = "full_sample",
) -> tuple[dict[str, list[tuple[int | str, pd.DataFrame]]], list[str]]:
    tables_by_spec: dict[str, list[tuple[int | str, pd.DataFrame]]] = {}
    skipped: list[str] = []

    for weight_spec in weight_specs:
        spec_tables: list[tuple[int | str, pd.DataFrame]] = []
        for aum_millions in aums:
            try:
                table = load_prediction_quantile_table(
                    analysis_dir=analysis_dir,
                    model=model,
                    weight_spec=weight_spec,
                    portfolio_run=portfolio_run,
                    aum_millions=aum_millions,
                    stock_universe=stock_universe,
                )
            except (FileNotFoundError, ValueError) as exc:
                if not skip_missing:
                    raise
                skipped.append(
                    f"{model}/{weight_spec}/{portfolio_run}/{stock_universe}/"
                    f"{_aum_label(aum_millions)}: {exc}"
                )
                continue
            spec_tables.append((aum_millions, table))
        if spec_tables:
            tables_by_spec[weight_spec] = spec_tables

    return tables_by_spec, skipped


def _load_model_spec_quantile_decomp_tables(
    analysis_dir: Path,
    model: str,
    weight_specs: Sequence[str],
    portfolio_run: str,
    aums: Sequence[int | str],
    skip_missing: bool,
    stock_universe: str = "full_sample",
) -> tuple[dict[str, list[tuple[int | str, dict[int, dict[str, float]]]]], list[str]]:
    tables_by_spec: dict[str, list[tuple[int | str, dict[int, dict[str, float]]]]] = {}
    skipped: list[str] = []

    for weight_spec in weight_specs:
        spec_tables: list[tuple[int | str, dict[int, dict[str, float]]]] = []
        for aum_millions in aums:
            try:
                table = load_prediction_quantile_decomposition_tables(
                    analysis_dir=analysis_dir,
                    model=model,
                    weight_spec=weight_spec,
                    portfolio_run=portfolio_run,
                    aum_millions=aum_millions,
                    stock_universe=stock_universe,
                )
            except (FileNotFoundError, ValueError) as exc:
                if not skip_missing:
                    raise
                skipped.append(
                    f"{model}/{weight_spec}/{portfolio_run}/{stock_universe}/"
                    f"{_aum_label(aum_millions)}: {exc}"
                )
                continue
            spec_tables.append((aum_millions, table))
        if spec_tables:
            tables_by_spec[weight_spec] = spec_tables

    return tables_by_spec, skipped


def write_table14_excel(
    tables_by_quantile: dict[int, dict[str, float]],
    output_path: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    aum_millions: int | str,
    stock_universe: str = "full_sample",
) -> None:
    """Write one single-spec Table 14 workbook with Q1-Q5 2x3 blocks."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 14"
    ws.sheet_view.showGridLines = False
    _set_table12_widths(ws)

    ws.merge_cells("A1:D1")
    ws["A1"] = (
        "Table 14: Standalone Prediction-Quantile 2x3 Decompositions - "
        f"{_model_title(model)} - {_spec_title(weight_spec)}"
    )
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = (
        f"Portfolio run: {_portfolio_run_title(portfolio_run)}. "
        f"Stock universe: {_stock_universe_title(stock_universe)}. "
        f"AUM case: {_aum_title(aum_millions)}."
    )
    ws["A2"].font = Font(size=9, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    next_row = 4
    for quantile, table in sorted(tables_by_quantile.items()):
        title = f"Standalone Q{quantile} 2x3 Decomposition ({_aum_title(aum_millions)})"
        next_row = _write_table12_block(
            ws=ws,
            table=table,
            start_row=next_row,
            title=title,
        )
        next_row += 1

    for row_idx in range(1, next_row):
        ws.row_dimensions[row_idx].height = 20
    ws.freeze_panes = "A3"

    wb.save(output_path)


def write_table14_model_workbook(
    model: str,
    tables_by_spec: dict[str, list[tuple[int | str, dict[int, dict[str, float]]]]],
    output_path: Path,
    portfolio_run: str,
    stock_universe: str = "full_sample",
    skipped: Sequence[str] | None = None,
) -> None:
    """Write one model workbook with Q1-Q5 Table-12-style blocks per spec."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active

    for idx, (weight_spec, spec_tables) in enumerate(tables_by_spec.items()):
        ws = default_ws if idx == 0 else wb.create_sheet()
        ws.title = _safe_spec_sheet_title(weight_spec)
        ws.sheet_view.showGridLines = False
        _set_table12_widths(ws)

        ws.merge_cells("A1:D1")
        ws["A1"] = (
            "Standalone Prediction-Quantile 2x3 Decompositions - "
            f"{_model_title(model)} - {_spec_title(weight_spec)}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = (
            f"Portfolio run: {_portfolio_run_title(portfolio_run)}. "
            f"Stock universe: {_stock_universe_title(stock_universe)}. "
            "Each AUM block reports separate Q1-Q5 2x3 decompositions."
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

        next_row = 4
        for aum_millions, quantile_tables in spec_tables:
            for quantile, table in sorted(quantile_tables.items()):
                title = (
                    f"Standalone Q{quantile} 2x3 Decomposition "
                    f"({_aum_title(aum_millions)})"
                )
                next_row = _write_table12_block(
                    ws=ws,
                    table=table,
                    start_row=next_row,
                    title=title,
                )
                next_row += 1
            next_row += 1

        for row_idx in range(1, next_row):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A3"

    if skipped:
        ws = wb.create_sheet("Skipped")
        ws["A1"] = "Skipped combinations"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = (
            "These combinations did not have complete 21e prediction-quantile "
            "outputs."
        )
        for row_idx, message in enumerate(skipped, start=4):
            ws.cell(row=row_idx, column=1, value=message)
        ws.column_dimensions["A"].width = 120

    wb.save(output_path)


def _write_table13_block(
    ws,
    table: pd.DataFrame,
    start_row: int,
    title: str,
    subtitle: str | None = None,
) -> int:
    """Write one standalone prediction-quantile portfolio block."""
    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="404040")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=11)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = title_row + 1
    if subtitle:
        subtitle_row = title_row + 1
        ws.merge_cells(
            start_row=subtitle_row,
            start_column=1,
            end_row=subtitle_row,
            end_column=11,
        )
        subtitle_cell = ws.cell(row=subtitle_row, column=1, value=subtitle)
        subtitle_cell.font = Font(italic=True, size=10, color="555555")
        subtitle_cell.alignment = Alignment(horizontal="center")
        header_row = title_row + 2

    headers = [
        "Cell",
        "Training",
        "Sort",
        "Quantile",
        "Gross ret (%)",
        "TC (%)",
        "Net ret (%)",
        "Gross SR",
        "Net SR",
        "Turnover",
        "Avg N",
    ]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=medium)
        cell.fill = header_fill

    data_start = header_row + 1
    for row_offset, (_, row) in enumerate(table.iterrows()):
        excel_row = data_start + row_offset
        values = [
            row["cell"],
            row["training_model"],
            row["portfolio_sort"],
            row["prediction_quantile"],
            row["gross_return_pct"],
            row["transaction_cost_pct"],
            row["net_return_pct"],
            row["gross_sr"],
            row["net_sr"],
            row["turnover"],
            row["avg_n_stocks"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in {2, 3} else "center"
            )
            if col_idx in {5, 6, 7, 8, 9, 10}:
                cell.number_format = "0.000"
            if col_idx == 11:
                cell.number_format = "0"
            if row_offset % 5 == 0:
                cell.border = Border(top=thin)

    bottom_row = data_start + len(table) - 1
    for col_idx in range(1, 12):
        ws.cell(row=bottom_row, column=col_idx).border = Border(bottom=medium)

    return bottom_row + 1


def write_table13_excel(
    table: pd.DataFrame,
    output_path: Path,
    model: str,
    weight_spec: str,
    portfolio_run: str,
    aum_millions: int | str,
    stock_universe: str = "full_sample",
) -> None:
    """Write a formatted single-AUM Table 13 workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table 13"
    ws.sheet_view.showGridLines = False
    _set_table13_widths(ws)

    title = (
        "Table 13: Standalone Prediction-Quantile Portfolios "
        f"({_aum_title(aum_millions)})"
    )
    subtitle = (
        f"{_model_title(model)} - {_spec_title(weight_spec)} - "
        f"{_portfolio_run_title(portfolio_run)} - "
        f"{_stock_universe_title(stock_universe)}"
    )
    next_row = _write_table13_block(
        ws=ws,
        table=table,
        start_row=1,
        title=title,
        subtitle=subtitle,
    )
    for row_idx in range(1, next_row):
        ws.row_dimensions[row_idx].height = 20
    ws.freeze_panes = "A4"

    wb.save(output_path)


def write_table13_model_workbook(
    model: str,
    tables_by_spec: dict[str, list[tuple[int | str, pd.DataFrame]]],
    output_path: Path,
    portfolio_run: str,
    stock_universe: str = "full_sample",
    skipped: Sequence[str] | None = None,
) -> None:
    """Write one model workbook with one prediction-quantile sheet per spec."""
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    default_ws = wb.active

    for idx, (weight_spec, spec_tables) in enumerate(tables_by_spec.items()):
        ws = default_ws if idx == 0 else wb.create_sheet()
        ws.title = _safe_spec_sheet_title(weight_spec)
        ws.sheet_view.showGridLines = False
        _set_table13_widths(ws)

        ws.merge_cells("A1:K1")
        ws["A1"] = (
            f"Standalone Prediction-Quantile Portfolios - {_model_title(model)} - "
            f"{_spec_title(weight_spec)}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:K2")
        ws["A2"] = (
            f"Portfolio run: {_portfolio_run_title(portfolio_run)}. "
            f"Stock universe: {_stock_universe_title(stock_universe)}. "
            "Each block reports one AUM case."
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

        next_row = 4
        for aum_millions, table in spec_tables:
            title = (
                "Standalone Prediction-Quantile Portfolios "
                f"({_aum_title(aum_millions)})"
            )
            next_row = _write_table13_block(
                ws=ws,
                table=table,
                start_row=next_row,
                title=title,
            )
            next_row += 2

        for row_idx in range(1, next_row):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A4"

    if skipped:
        ws = wb.create_sheet("Skipped")
        ws["A1"] = "Skipped combinations"
        ws["A1"].font = Font(bold=True)
        ws["A2"] = (
            "These combinations did not have complete 21e prediction-quantile "
            "outputs."
        )
        for row_idx, message in enumerate(skipped, start=4):
            ws.cell(row=row_idx, column=1, value=message)
        ws.column_dimensions["A"].width = 120

    wb.save(output_path)


def _table_output_dir(output_dir: Path, model: str, stock_universe: str) -> Path:
    """Return the organized report-table directory for one model and universe."""
    return output_dir / model / stock_universe


def _default_output_path(
    output_dir: Path,
    table_name: str,
    model: str,
    weight_specs: Sequence[str],
    portfolio_run: str,
    aums: Sequence[int | str],
    stock_universe: str = "full_sample",
) -> Path:
    spec_part = _selection_label(
        weight_specs,
        WEIGHT_SPEC_CHOICES,
        "all_specs",
        "specs",
    )
    aum_part = _selection_label(aums, DEFAULT_AUMS, "all_aums", "aums")
    safe_name = (
        f"{table_name}_{spec_part}_{portfolio_run}_{aum_part}.xlsx"
    )
    return _table_output_dir(output_dir, model, stock_universe) / safe_name


def _resolve_model_output_path(
    output_arg: str | None,
    default_dir: Path,
    table_name: str,
    model: str,
    weight_specs: Sequence[str],
    portfolio_run: str,
    aums: Sequence[int | str],
    n_workbooks: int,
    stock_universe: str = "full_sample",
) -> Path:
    if not output_arg:
        return _default_output_path(
            output_dir=default_dir,
            table_name=table_name,
            model=model,
            weight_specs=weight_specs,
            portfolio_run=portfolio_run,
            aums=aums,
            stock_universe=stock_universe,
        )

    requested = Path(output_arg)
    if n_workbooks == 1 and requested.suffix == ".xlsx":
        return requested

    output_dir = requested if requested.suffix != ".xlsx" else requested.parent
    return _default_output_path(
        output_dir=output_dir,
        table_name=table_name,
        model=model,
        weight_specs=weight_specs,
        portfolio_run=portfolio_run,
        aums=aums,
        stock_universe=stock_universe,
    )


def main() -> None:
    args = parse_args()
    config = load_config()
    analysis_dir = Path(config["project"]["output_dir"]) / "formalanalysis" / "analysis"
    output_dir = Path(config["project"]["output_dir"]) / "formalanalysis" / "tables"

    models = _parse_selection(args.model, MODEL_CHOICES, "model")
    weight_specs = _parse_weight_specs(args.weight_spec)
    portfolio_runs = _parse_portfolio_runs(args.portfolio_run)
    aums = _parse_aums(args.aum)
    stock_universes = resolve_stock_universes(args.stock_universe)

    single_table = (
        len(models)
        == len(weight_specs)
        == len(portfolio_runs)
        == len(aums)
        == len(stock_universes)
        == 1
    )
    if single_table:
        portfolio_run = portfolio_runs[0]
        stock_universe = stock_universes[0]
        if args.output:
            output_path = Path(args.output)
        else:
            output_path = _default_output_path(
                output_dir=output_dir,
                table_name=args.table,
                model=models[0],
                weight_specs=weight_specs,
                portfolio_run=portfolio_run,
                aums=aums,
                stock_universe=stock_universe,
            )
        if args.table == "table13":
            table = load_prediction_quantile_table(
                analysis_dir=analysis_dir,
                model=models[0],
                weight_spec=weight_specs[0],
                portfolio_run=portfolio_run,
                aum_millions=aums[0],
                stock_universe=stock_universe,
            )
            write_table13_excel(
                table=table,
                output_path=output_path,
                model=models[0],
                weight_spec=weight_specs[0],
                portfolio_run=portfolio_run,
                aum_millions=aums[0],
                stock_universe=stock_universe,
            )
        else:
            table = load_prediction_quantile_decomposition_tables(
                analysis_dir=analysis_dir,
                model=models[0],
                weight_spec=weight_specs[0],
                portfolio_run=portfolio_run,
                aum_millions=aums[0],
                stock_universe=stock_universe,
            )
            write_table14_excel(
                tables_by_quantile=table,
                output_path=output_path,
                model=models[0],
                weight_spec=weight_specs[0],
                portfolio_run=portfolio_run,
                aum_millions=aums[0],
                stock_universe=stock_universe,
            )
        print(output_path)
        return

    output_paths: list[Path] = []
    skipped_runs: list[str] = []
    for model in models:
        for portfolio_run in portfolio_runs:
            for stock_universe in stock_universes:
                if args.table == "table13":
                    tables_by_spec, skipped = _load_model_spec_quantile_tables(
                        analysis_dir=analysis_dir,
                        model=model,
                        weight_specs=weight_specs,
                        portfolio_run=portfolio_run,
                        aums=aums,
                        skip_missing=args.skip_missing,
                        stock_universe=stock_universe,
                    )
                else:
                    tables_by_spec, skipped = _load_model_spec_quantile_decomp_tables(
                        analysis_dir=analysis_dir,
                        model=model,
                        weight_specs=weight_specs,
                        portfolio_run=portfolio_run,
                        aums=aums,
                        skip_missing=args.skip_missing,
                        stock_universe=stock_universe,
                    )
                if not tables_by_spec:
                    if args.skip_missing:
                        skipped_runs.append(f"{model}/{portfolio_run}/{stock_universe}")
                        continue
                    raise FileNotFoundError(
                        "No matching 21e outputs were available for "
                        f"{model}/{portfolio_run}/{stock_universe}."
                    )

                output_path = _resolve_model_output_path(
                    output_arg=args.output,
                    default_dir=output_dir,
                    table_name=args.table,
                    model=model,
                    weight_specs=weight_specs,
                    portfolio_run=portfolio_run,
                    aums=aums,
                    n_workbooks=(
                        len(models) * len(portfolio_runs) * len(stock_universes)
                    ),
                    stock_universe=stock_universe,
                )
                if args.table == "table13":
                    write_table13_model_workbook(
                        model=model,
                        tables_by_spec=tables_by_spec,
                        output_path=output_path,
                        portfolio_run=portfolio_run,
                        stock_universe=stock_universe,
                        skipped=skipped,
                    )
                else:
                    write_table14_model_workbook(
                        model=model,
                        tables_by_spec=tables_by_spec,
                        output_path=output_path,
                        portfolio_run=portfolio_run,
                        stock_universe=stock_universe,
                        skipped=skipped,
                    )
                output_paths.append(output_path)

    if not output_paths:
        skipped = ", ".join(skipped_runs)
        raise FileNotFoundError(f"No workbooks were written. Skipped runs: {skipped}")

    for output_path in output_paths:
        print(output_path)


if __name__ == "__main__":
    main()
