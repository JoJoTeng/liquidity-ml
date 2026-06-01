"""Reconstruct Table 12 with a two-sided TC-aware long/short sort.

This script does NOT rerun 21e and does NOT retrain anything. It reads the saved
per-stock predictions (standard, weighted, tc_target standard, tc_target
weighted) plus the processed panel, then rebuilds the six 2x3 cells using the
*two-sided* sort already implemented in ``build_long_short_portfolio``:

    Cell  Predictions           Sort
    1A    standard              plain: long top of r_hat, short bottom of r_hat
    1B    standard              tc_penalised: long top of (r_hat - TC),
                                              short bottom of (r_hat + TC)
    1C    tc_target standard    tc_target_score: long top of s,
                                                 short bottom of (s + 2*TC)
    2A/2B/2C  weighted / tc_target weighted -- same three sorts

This differs from the current ``compute_two_by_three_decomposition``, which
ranks every stock by a single score and takes Q_high - Q_low (so its short leg
is just the bottom of the same score). Here the long and short legs use
different signals, which is the economically clean TC-aware construction.

Capital convention (``--leg-capital``): ``full`` (default) trades the full AUM
on each leg, matching script 22's quantile path, so net Sharpes are comparable
to the single-score Table 12; ``half`` splits AUM/2 across the two legs (the
native long-short default).

Outputs are written under
``outputs/formalanalysis/tables/{model}/{stock_universe}/`` by default, so
``--stock-universe both`` writes separate ``full_sample`` and ``nyse``
directories.

  - ``table12_two_sided_{portfolio_weighting}.xlsx``
    -- one workbook per model/portfolio-weighting, one sheet per weight spec,
    one block per AUM, in the Table 12 layout. If ``--leg-capital half`` is
    used, ``halfleg`` is inserted after ``two_sided``.
  - ``table13_legs_two_sided_{portfolio_weighting}.xlsx`` -- formatted
    Table-13-style per-leg breakdown: for every cell (1A-2C) the Long leg,
    Short leg, and Long-Short, each with gross/net return, gross/net SR, TC,
    turnover, and avg stock count.
  - ``table14_legdecomp_two_sided_{portfolio_weighting}.xlsx`` -- formatted
    Table-14 analog: a 2x3 net-Sharpe decomposition (six cells plus
    training/portfolio/TC-target effects) computed within each of the Long leg,
    Short leg, and Long-Short.
"""

from __future__ import annotations

import argparse
import importlib.util
import logging
import os
import sys
import tempfile
from pathlib import Path

import numpy as np
import pandas as pd

_runtime_cache = Path(tempfile.gettempdir()) / "liquidity_ml_cache"
_runtime_cache.mkdir(parents=True, exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_runtime_cache / "matplotlib"))
os.environ.setdefault("XDG_CACHE_HOME", str(_runtime_cache / "xdg"))

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.analysis.formal.common import is_proportional_tc_scenario  # noqa: E402
from src.analysis.formal.script_utils import (  # noqa: E402
    add_stock_universe_option,
    formal_output_dirs,
    resolve_stock_universes,
)
from src.config import load_config  # noqa: E402
from src.data.loader import load_processed_panel  # noqa: E402
from src.evaluation.statistics import sharpe_ratio  # noqa: E402
from src.portfolio.construction import (  # noqa: E402
    _compute_net_returns_with_context,
    build_portfolio_timeseries,
    prepare_transaction_cost_context,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("table12_two_sided")

MODEL_CHOICES = ["elastic_net", "xgboost", "neural_network"]
PRIMARY_WEIGHT_SPECS = ["dolvol", "softmax_rank_lam2", "tc_500m", "tc_rank_lam3_500m"]
DEFAULT_AUMS = ["proportional_tc", 100, 500, 1000]
CELLS = ["1A", "1B", "1C", "2A", "2B", "2C"]

# (cell, prediction key, tc_penalised, tc_target_score)
# Standard cells use the model's standard / tc_target-standard predictions,
# which are shared across all weight specs, so they are built and netted once
# per model. Weighted cells depend on the weight spec.
STANDARD_CELL_SPECS = [
    ("1A", "standard", False, False),
    ("1B", "standard", True, False),
    ("1C", "tc_target_standard", False, True),
]
WEIGHTED_CELL_SPECS = [
    ("2A", "weighted", False, False),
    ("2B", "weighted", True, False),
    ("2C", "tc_target_weighted", False, True),
]


def _load_script22():
    """Import the numerically-named script 22 module for its Table 12 writers."""
    path = Path(__file__).resolve().parent / "22_prepare_portfolio_excel_tables.py"
    spec = importlib.util.spec_from_file_location("prepare_portfolio_excel_tables", path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def parse_args():
    parser = argparse.ArgumentParser(
        description="Reconstruct Table 12 with a two-sided TC-aware long/short sort."
    )
    parser.add_argument(
        "--model",
        default="all",
        help="Model(s): one, comma-separated, or 'all'. Default all.",
    )
    parser.add_argument(
        "--weight-spec",
        default="primary",
        help="Weight spec(s): one, comma-separated, or 'primary'. Default primary.",
    )
    parser.add_argument(
        "--aum",
        default="all",
        help=(
            "AUM(s) in $M, proportional_tc, comma-separated, or 'all'. "
            "Default all = proportional_tc,100,500,1000."
        ),
    )
    parser.add_argument(
        "--portfolio-weighting",
        default="equal",
        choices=["equal", "value", "both"],
        help=(
            "Selected-leg weighting within each leg. Use 'both' to write "
            "separate equal and value-weighted workbooks. Default equal."
        ),
    )
    add_stock_universe_option(parser)
    parser.add_argument(
        "--leg-capital",
        default="full",
        choices=["full", "half"],
        help=(
            "Per-leg capital. 'full' trades the full AUM on each leg (matches "
            "script 22's convention); 'half' splits AUM/2 across legs. Default full."
        ),
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help=(
            "Output root directory. Default "
            "outputs/formalanalysis/tables/{model}/{stock_universe}/."
        ),
    )
    return parser.parse_args()


def _resolve_models(value: str) -> list[str]:
    if value.lower() == "all":
        return list(MODEL_CHOICES)
    return [m.strip() for m in value.split(",") if m.strip()]


def _resolve_specs(value: str) -> list[str]:
    if value.lower() == "primary":
        return list(PRIMARY_WEIGHT_SPECS)
    return [s.strip() for s in value.split(",") if s.strip()]


def _resolve_aums(value: str) -> list[int | str]:
    """Resolve AUM CLI values to $M integers plus proportional_tc aliases."""
    if value.lower() == "all":
        return list(DEFAULT_AUMS)

    out: list[int | str] = []
    for raw in value.split(","):
        token = raw.strip()
        if not token:
            continue
        if is_proportional_tc_scenario(token):
            out.append("proportional_tc")
        else:
            out.append(int(token))
    if not out:
        raise ValueError("No AUM values selected.")
    return out


def _resolve_portfolio_weightings(value: str) -> list[str]:
    """Resolve equal/value/both to concrete portfolio-weighting values."""
    if value == "both":
        return ["equal", "value"]
    return [value]


def _filter_panel_for_stock_universe(
    panel: pd.DataFrame,
    stock_universe: str,
) -> pd.DataFrame:
    """Return the panel rows used for one stock-universe portfolio run."""
    if stock_universe == "full_sample":
        return panel
    if stock_universe != "nyse":
        raise ValueError(f"Unknown stock universe: {stock_universe!r}")
    if "exchcd" not in panel.columns:
        raise ValueError("NYSE stock-universe filter requires panel column 'exchcd'.")
    return panel.loc[panel["exchcd"] == 1].copy()


def _leg_capital_filename_part(leg_capital: str) -> str:
    """Return optional filename token for non-default leg-capital settings."""
    if leg_capital == "full":
        return ""
    return f"{leg_capital}leg_"


def _table_output_dir(output_dir: Path, model: str, stock_universe: str) -> Path:
    """Return the organized report-table directory for one model and universe."""
    return output_dir / model / stock_universe


def _stock_universe_title(mod22, stock_universe: str) -> str:
    """Return the script-22 stock-universe display label."""
    return mod22._stock_universe_title(stock_universe)


def _aum_title(mod22, aum_case: int | str) -> str:
    """Return script-22-compatible title text for an AUM case."""
    return mod22._aum_title(aum_case)


def _standard_prediction_paths(experiment_dir: Path, model: str) -> dict[str, Path]:
    """Prediction files for the standard cells (shared across a model's specs)."""
    model_dir = experiment_dir / model
    return {
        "standard": model_dir / "standard" / "predictions.parquet",
        "tc_target_standard": model_dir / "tc_target" / "standard" / "predictions.parquet",
    }


def _weighted_prediction_paths(
    experiment_dir: Path, model: str, weight_spec: str,
) -> dict[str, Path]:
    """Prediction files for the weighted cells of one weight spec."""
    model_dir = experiment_dir / model
    return {
        "weighted": model_dir / weight_spec / "predictions.parquet",
        "tc_target_weighted": model_dir / "tc_target" / weight_spec / "predictions.parquet",
    }


def _build_cell_gross(panel, preds, tc_penalised, tc_target_score, build_aum, config, weighting):
    """Build one cell's gross long-short time series + positions (AUM-independent).

    The TC-aware sort uses half-spread only, so gross returns and positions do
    not depend on AUM; only the later net-return step does.
    """
    ret_df, pos_hist = build_portfolio_timeseries(
        panel,
        preds,
        tc_penalised=tc_penalised,
        tc_target_score=tc_target_score,
        aum=build_aum if (tc_penalised or tc_target_score) else None,
        config=config,
        portfolio_mode="long_short",
        portfolio_weighting=weighting,
    )
    return ret_df, pos_hist


def _decomposition_dict(net_by_cell: dict[str, pd.DataFrame]) -> dict[str, float]:
    """Annualized net-Sharpe Table 12 dict, mirroring script 22's schema."""
    common = set(net_by_cell["1A"]["yyyymm"])
    for cell in CELLS[1:]:
        common &= set(net_by_cell[cell]["yyyymm"])
    common = sorted(common)
    if not common:
        raise ValueError("No common months across the six cells")

    net_sr = {}
    for cell in CELLS:
        df = net_by_cell[cell]
        df = df[df["yyyymm"].isin(common)].sort_values("yyyymm")
        net_sr[cell] = sharpe_ratio(df["ret_long_short_net"].values, annualize=True)

    training_effect = net_sr["2A"] - net_sr["1A"]
    portfolio_effect = net_sr["1B"] - net_sr["1A"]
    total_effect = net_sr["2B"] - net_sr["1A"]
    interaction = total_effect - training_effect - portfolio_effect
    return {
        "SR_net_1A": net_sr["1A"],
        "SR_net_1B": net_sr["1B"],
        "SR_net_1C": net_sr["1C"],
        "SR_net_2A": net_sr["2A"],
        "SR_net_2B": net_sr["2B"],
        "SR_net_2C": net_sr["2C"],
        "training_effect": training_effect,
        "portfolio_effect": portfolio_effect,
        "total_effect": total_effect,
        "interaction": interaction,
        "training_share_pct": (
            training_effect / total_effect * 100.0
            if abs(total_effect) > 1e-10
            else float("nan")
        ),
        "tc_target_standard_vs_prediction": net_sr["1C"] - net_sr["1A"],
        "tc_target_standard_vs_tc_sort": net_sr["1C"] - net_sr["1B"],
        "tc_target_weighted_vs_prediction": net_sr["2C"] - net_sr["2A"],
        "tc_target_weighted_vs_tc_sort": net_sr["2C"] - net_sr["2B"],
        "tc_target_training_effect": net_sr["2C"] - net_sr["1C"],
        "tc_target_total_effect": net_sr["2C"] - net_sr["1A"],
    }


# Cell -> (training-model code, portfolio-sort code) labels, matching script 22.
CELL_META = {
    "1A": ("standard", "prediction"),
    "1B": ("standard", "tc_net_score"),
    "1C": ("standard_tc_target", "tc_target_score"),
    "2A": ("weighted", "prediction"),
    "2B": ("weighted", "tc_net_score"),
    "2C": ("weighted_tc_target", "tc_target_score"),
}


def _summarize_leg(net: pd.DataFrame, n_series) -> dict[str, float]:
    """Table-13-style annualized summary of one net-return series."""
    turnover = float(net["turnover"].iloc[1:].mean()) if len(net) > 1 else float("nan")
    return {
        "gross_return_pct": float(net["ret_long_short"].mean() * 100.0),
        "transaction_cost_pct": float(net["transaction_cost"].mean() * 100.0),
        "net_return_pct": float(net["ret_long_short_net"].mean() * 100.0),
        "gross_sr": sharpe_ratio(net["ret_long_short"].values, annualize=True),
        "net_sr": sharpe_ratio(net["ret_long_short_net"].values, annualize=True),
        "turnover": turnover,
        "avg_n_stocks": float(np.asarray(n_series, dtype=float).mean()),
    }


def _leg_net(
    ret_df,
    pos_hist,
    ret_col,
    leg_key,
    aum,
    tc_context,
    leg_aum_override,
    proportional_tc_only,
):
    """Net returns for one leg treated as a standalone long-only book.

    This mirrors how Table 13 treats each prediction quantile: the leg's own
    positions are the book, its raw weighted return is the gross return, and TC
    is computed on that leg's turnover alone.
    """
    pos = {ym: {"long": pos_hist[ym][leg_key], "short": {}} for ym in pos_hist}
    gross = ret_df[["yyyymm", ret_col]].rename(columns={ret_col: "ret_long_short"})
    return _compute_net_returns_with_context(
        gross, pos, aum=aum, tc_context=tc_context,
        portfolio_mode="long_short", leg_aum_override=leg_aum_override,
        proportional_tc_only=proportional_tc_only,
    )


def _compute_cell_nets(
    gross_cells,
    aum,
    tc_context,
    leg_aum_override,
    proportional_tc_only,
) -> dict:
    """Per cell, the net returns for the long leg, short leg, and long-short.

    Computed once here and shared by Table 12, Table 13, and Table 14 so each
    cell's three net-return series are not recomputed three times over. Iterates
    over whatever cells are in ``gross_cells`` (standard or weighted subset).
    """
    cell_nets = {}
    for cell, (ret_df, pos_hist) in gross_cells.items():
        cell_nets[cell] = {
            "long": _leg_net(
                ret_df, pos_hist, "ret_long", "long", aum, tc_context,
                leg_aum_override, proportional_tc_only,
            ),
            "short": _leg_net(
                ret_df, pos_hist, "ret_short", "short", aum, tc_context,
                leg_aum_override, proportional_tc_only,
            ),
            "ls": _compute_net_returns_with_context(
                ret_df[["yyyymm", "ret_long_short"]].copy(), pos_hist, aum=aum,
                tc_context=tc_context, portfolio_mode="long_short",
                leg_aum_override=leg_aum_override,
                proportional_tc_only=proportional_tc_only,
            ),
            "n_long": ret_df["n_long"],
            "n_short": ret_df["n_short"],
        }
    return cell_nets


def _leg_breakdown_rows(model, weight_spec, aum_m, leg_capital, cell_nets) -> list[dict]:
    """Per-cell Long-leg / Short-leg / Long-Short rows (Table-13 style).

    The short row is reported as the economic short-position contribution:
    gross = -R_Q1 and net = -R_Q1 - TC_Q1.
    """
    rows = []
    for cell in CELLS:
        training, sort = CELL_META[cell]
        cn = cell_nets[cell]
        legs = [
            ("Long (Q5)", cn["long"], cn["n_long"]),
            ("Short (Q1)", _short_side_contribution(cn["short"]), cn["n_short"]),
            ("Long-Short", cn["ls"], cn["n_long"] + cn["n_short"]),
        ]
        for leg_label, net, n_series in legs:
            rows.append({
                "model": model, "weight_spec": weight_spec, "aum_m": aum_m,
                "leg_capital": leg_capital, "cell": cell,
                "training_model": training, "portfolio_sort": sort, "leg": leg_label,
                **_summarize_leg(net, n_series),
            })
    return rows


def _leg_decomposition_rows(model, weight_spec, aum_m, leg_capital, cell_nets) -> list[dict]:
    """Table-14 analog: a Table-12-style 2x3 decomposition within each leg.

    For the Long leg, Short leg, and Long-Short, collect the six cells' net
    returns and run the same annualized-net-Sharpe decomposition as Table 12.
    This attributes the training / portfolio / TC-target effects to each side.
    """
    rows = []
    leg_specs = [
        ("Long", "long", False),
        ("Short", "short", True),
        ("Long-Short", "ls", False),
    ]
    for leg_label, key, short_contribution in leg_specs:
        if short_contribution:
            net_by_cell = {
                cell: _short_side_contribution(cell_nets[cell][key])
                for cell in CELLS
            }
        else:
            net_by_cell = {cell: cell_nets[cell][key] for cell in CELLS}
        table = _decomposition_dict(net_by_cell)
        rows.append({
            "model": model, "weight_spec": weight_spec, "aum_m": aum_m,
            "leg_capital": leg_capital, "leg": leg_label, **table,
        })
    return rows


def _short_side_contribution(net: pd.DataFrame) -> pd.DataFrame:
    """Return the economic short-side contribution series for Table 14.

    Table 13 intentionally reports the raw return of the Q1 stock basket. Table
    14 is a contribution decomposition, so the short side should enter with the
    sign used by the long-short portfolio:

        gross_short_contribution = -r_Q1
        net_short_contribution = -r_Q1 - TC_Q1
    """
    out = net.copy()
    gross = -out["ret_long_short"]
    out["ret_long_short"] = gross
    out["ret_long_short_net"] = gross - out["transaction_cost"]
    return out


def _write_table12_workbook(
    mod22, model, tables_by_spec, output_path, weighting, leg_capital,
    stock_universe,
) -> None:
    """Write a two-sided Table 12 workbook (one sheet per spec, block per AUM).

    Reuses script 22's kept 2x3 layout primitives (``_write_table12_block`` /
    ``_set_table12_widths`` and the label helpers) so the formatting matches the
    paper tables, without depending on script 22's removed Table 12 report code.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_ws = wb.active
    for idx, (weight_spec, spec_tables) in enumerate(tables_by_spec.items()):
        ws = default_ws if idx == 0 else wb.create_sheet()
        ws.title = mod22._safe_spec_sheet_title(weight_spec)
        ws.sheet_view.showGridLines = False
        mod22._set_table12_widths(ws)

        ws.merge_cells("A1:D1")
        ws["A1"] = (
            f"Two-Sided 2x3 Q5-Q1 Decomposition - "
            f"{mod22._model_title(model)} - {mod22._spec_title(weight_spec)}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = (
            f"Two-sided TC-aware long/short. Selected-leg weighting: {weighting}. "
            f"Per-leg capital: {leg_capital}. "
            f"Stock universe: {_stock_universe_title(mod22, stock_universe)}. "
            "Each block reports one AUM case."
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

        next_row = 4
        for aum_millions, table in spec_tables:
            title = (
                f"Two-Sided 2x3 Q5-Q1 Decomposition ({mod22._aum_title(aum_millions)})"
            )
            next_row = mod22._write_table12_block(
                ws=ws, table=table, start_row=next_row, title=title,
            )
            next_row += 1

        for row_idx in range(1, next_row):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A3"

    wb.save(output_path)


def _format_table13_leg_rows_for_excel(rows: list[dict]) -> pd.DataFrame:
    """Return rows in the schema expected by the formatted Table 13 block."""
    df = pd.DataFrame(rows)
    if df.empty:
        return df

    cell_order = {cell: idx for idx, cell in enumerate(CELLS)}
    leg_order = {"Long (Q5)": 0, "Short (Q1)": 1, "Long-Short": 2}
    df["_cell_order"] = df["cell"].map(cell_order).fillna(99)
    df["_leg_order"] = df["leg"].map(leg_order).fillna(99)
    df = df.sort_values(["_cell_order", "_leg_order"]).drop(
        columns=["_cell_order", "_leg_order"]
    )
    out = df[
        [
            "cell",
            "training_model",
            "portfolio_sort",
            "leg",
            "gross_return_pct",
            "transaction_cost_pct",
            "net_return_pct",
            "gross_sr",
            "net_sr",
            "turnover",
            "avg_n_stocks",
        ]
    ].copy()
    out["training_model"] = out["training_model"].map({
        "standard": "Standard",
        "weighted": "Weighted",
        "standard_tc_target": "Standard TC-target",
        "weighted_tc_target": "Weighted TC-target",
    }).fillna(out["training_model"])
    out["portfolio_sort"] = out["portfolio_sort"].map({
        "prediction": "Sort on r_hat",
        "tc_net_score": "TC-aware sort",
        "tc_target_score": "TC-target score",
    }).fillna(out["portfolio_sort"])
    return out


def _write_table13_legs_block(ws, table: pd.DataFrame, start_row: int, title: str) -> int:
    """Write one two-sided Table-13-style block with 3 leg rows per cell.

    Script 22's normal Table 13 writer separates rows in groups of five because
    it reports Q1-Q5 prediction-quantile portfolios. This two-sided table is
    different: each 2x3 cell has exactly three rows: Long, Short, Long-Short.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="404040")
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=11)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = title_row + 1
    headers = [
        "Cell",
        "Training",
        "Sort",
        "Leg",
        "Gross ret (%)",
        "TC (%)",
        "Net ret (%)",
        "Gross SR",
        "Net SR",
        "Turnover",
        "Avg N",
    ]
    for col_idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=medium)
        cell.fill = header_fill

    data_start = header_row + 1
    previous_cell = None
    for row_offset, (_, row) in enumerate(table.iterrows()):
        excel_row = data_start + row_offset
        values = [
            row["cell"],
            row["training_model"],
            row["portfolio_sort"],
            row["leg"],
            row["gross_return_pct"],
            row["transaction_cost_pct"],
            row["net_return_pct"],
            row["gross_sr"],
            row["net_sr"],
            row["turnover"],
            row["avg_n_stocks"],
        ]
        starts_new_cell = row["cell"] != previous_cell
        previous_cell = row["cell"]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="left" if col_idx in {2, 3, 4} else "center"
            )
            if col_idx in {5, 6, 7, 8, 9, 10}:
                cell.number_format = "0.000"
            if col_idx == 11:
                cell.number_format = "0"
            if starts_new_cell:
                cell.border = Border(top=thin)

    bottom_row = data_start + len(table) - 1
    for col_idx in range(1, 12):
        ws.cell(row=bottom_row, column=col_idx).border = Border(bottom=medium)

    return bottom_row + 1


def _write_table13_legs_workbook(
    mod22, model, rows, output_path, weighting, leg_capital, stock_universe,
) -> None:
    """Write two-sided leg summaries using script 22's Table 13 workbook style."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_ws = wb.active

    by_spec = pd.DataFrame(rows).groupby("weight_spec", sort=False)
    for idx, (weight_spec, spec_df) in enumerate(by_spec):
        ws = default_ws if idx == 0 else wb.create_sheet()
        ws.title = mod22._safe_spec_sheet_title(weight_spec)
        ws.sheet_view.showGridLines = False
        mod22._set_table13_widths(ws)

        ws.merge_cells("A1:K1")
        ws["A1"] = (
            f"Two-Sided Long/Short Leg Portfolios - {mod22._model_title(model)} - "
            f"{mod22._spec_title(weight_spec)}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:K2")
        ws["A2"] = (
            f"Selected-leg weighting: {weighting}. Per-leg capital: {leg_capital}. "
            f"Stock universe: {_stock_universe_title(mod22, stock_universe)}. "
            "Each block reports one AUM case."
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

        next_row = 4
        for aum_m, aum_df in spec_df.groupby("aum_m", sort=False):
            table = _format_table13_leg_rows_for_excel(aum_df.to_dict("records"))
            title = f"Two-Sided Long/Short Leg Portfolios ({_aum_title(mod22, aum_m)})"
            next_row = _write_table13_legs_block(
                ws=ws,
                table=table,
                start_row=next_row,
                title=title,
            )
            next_row += 2

        for row_idx in range(1, next_row):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A4"

    wb.save(output_path)


def _write_table14_legdecomp_workbook(
    mod22, model, rows, output_path, weighting, leg_capital, stock_universe,
) -> None:
    """Write leg-level 2x3 decompositions using script 22's Table 14 style."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default_ws = wb.active

    by_spec = pd.DataFrame(rows).groupby("weight_spec", sort=False)
    for idx, (weight_spec, spec_df) in enumerate(by_spec):
        ws = default_ws if idx == 0 else wb.create_sheet()
        ws.title = mod22._safe_spec_sheet_title(weight_spec)
        ws.sheet_view.showGridLines = False
        mod22._set_table12_widths(ws)

        ws.merge_cells("A1:D1")
        ws["A1"] = (
            f"Two-Sided Leg 2x3 Decompositions - {mod22._model_title(model)} - "
            f"{mod22._spec_title(weight_spec)}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:D2")
        ws["A2"] = (
            f"Selected-leg weighting: {weighting}. Per-leg capital: {leg_capital}. "
            f"Stock universe: {_stock_universe_title(mod22, stock_universe)}. "
            "Each AUM block reports Long, Short, and Long-Short decompositions."
        )
        ws["A2"].font = Font(size=9, color="555555")
        ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

        next_row = 4
        for aum_m, aum_df in spec_df.groupby("aum_m", sort=False):
            for leg, leg_df in aum_df.groupby("leg", sort=False):
                table = leg_df.iloc[0].to_dict()
                title = (
                    f"{leg} 2x3 Decomposition "
                    f"({_aum_title(mod22, aum_m)})"
                )
                next_row = mod22._write_table12_block(
                    ws=ws,
                    table=table,
                    start_row=next_row,
                    title=title,
                )
                next_row += 1
            next_row += 1

        for row_idx in range(1, next_row):
            ws.row_dimensions[row_idx].height = 20
        ws.freeze_panes = "A3"

    wb.save(output_path)


def main() -> None:
    args = parse_args()
    config = load_config()
    _, experiment_dir, _ = formal_output_dirs(config)
    mod22 = _load_script22()

    models = _resolve_models(args.model)
    weight_specs = _resolve_specs(args.weight_spec)
    aums = _resolve_aums(args.aum)
    portfolio_weightings = _resolve_portfolio_weightings(args.portfolio_weighting)
    stock_universes = resolve_stock_universes(args.stock_universe)
    leg_capital = args.leg_capital
    out_dir = Path(args.output_dir) if args.output_dir else (
        Path(config["project"]["output_dir"]) / "formalanalysis" / "tables"
    )
    out_dir.mkdir(parents=True, exist_ok=True)

    logger.info(
        (
            "Config: models=%s specs=%s aums=%s weightings=%s "
            "stock_universes=%s leg_capital=%s"
        ),
        models,
        weight_specs,
        aums,
        portfolio_weightings,
        stock_universes,
        leg_capital,
    )
    logger.info("Loading processed panel (this is the heavy step)...")
    panel_full = load_processed_panel()
    numeric_aums = [a for a in aums if not is_proportional_tc_scenario(a)]
    build_aum = (max(numeric_aums) if numeric_aums else 1) * 1_000_000

    written: list[Path] = []

    def leg_aum_for(aum_case):
        proportional_tc_only = is_proportional_tc_scenario(aum_case)
        aum = 0.0 if proportional_tc_only else float(aum_case) * 1_000_000
        # Capital convention. "full" gives each leg the full AUM
        # (leg_aum_override=aum), matching script 22's quantile path; "half"
        # splits AUM/2 across the two legs (the long-short default).
        leg_aum_override = (
            0.0
            if proportional_tc_only
            else (aum if leg_capital == "full" else None)
        )
        return aum, leg_aum_override, proportional_tc_only

    for stock_universe in stock_universes:
        panel = _filter_panel_for_stock_universe(panel_full, stock_universe)
        logger.info(
            "Stock universe %s: %d rows, %d months",
            stock_universe,
            len(panel),
            panel["yyyymm"].nunique(),
        )
        tc_context = prepare_transaction_cost_context(panel, config)
        leg_token = _leg_capital_filename_part(leg_capital)

        for weighting in portfolio_weightings:
            for model in models:
                model_out_dir = _table_output_dir(out_dir, model, stock_universe)
                # Standard cells (1A/1B/1C) use the model's shared standard
                # predictions, so build + net them once per model/universe and
                # reuse them across every weight spec.
                std_paths = _standard_prediction_paths(experiment_dir, model)
                std_missing = [str(p) for p in std_paths.values() if not p.exists()]
                if std_missing:
                    logger.warning(
                        "Skip model %s, missing standard predictions:\n  %s",
                        model,
                        "\n  ".join(std_missing),
                    )
                    continue
                std_preds = {k: pd.read_parquet(p) for k, p in std_paths.items()}

                logger.info(
                    "[%s/%s/%s] building 3 standard cells (shared across specs)",
                    stock_universe,
                    model,
                    weighting,
                )
                std_gross = {}
                for cell, pred_key, tc_pen, tc_tgt in STANDARD_CELL_SPECS:
                    std_gross[cell] = _build_cell_gross(
                        panel,
                        std_preds[pred_key],
                        tc_pen,
                        tc_tgt,
                        build_aum,
                        config,
                        weighting,
                    )
                std_nets_by_aum = {}
                for aum_case in aums:
                    aum, leg_aum_override, proportional_tc_only = leg_aum_for(aum_case)
                    std_nets_by_aum[aum_case] = _compute_cell_nets(
                        std_gross,
                        aum,
                        tc_context,
                        leg_aum_override,
                        proportional_tc_only,
                    )

                tables_by_spec: dict[str, list[tuple[int | str, dict]]] = {}
                model_leg_rows: list[dict] = []
                model_leg_decomp_rows: list[dict] = []
                for weight_spec in weight_specs:
                    wt_paths = _weighted_prediction_paths(
                        experiment_dir, model, weight_spec,
                    )
                    wt_missing = [str(p) for p in wt_paths.values() if not p.exists()]
                    if wt_missing:
                        logger.warning(
                            "Skip %s/%s, missing weighted predictions:\n  %s",
                            model,
                            weight_spec,
                            "\n  ".join(wt_missing),
                        )
                        continue
                    wt_preds = {k: pd.read_parquet(p) for k, p in wt_paths.items()}

                    logger.info(
                        "[%s/%s/%s/%s] building 3 weighted cells",
                        stock_universe,
                        model,
                        weighting,
                        weight_spec,
                    )
                    wt_gross = {}
                    for cell, pred_key, tc_pen, tc_tgt in WEIGHTED_CELL_SPECS:
                        wt_gross[cell] = _build_cell_gross(
                            panel,
                            wt_preds[pred_key],
                            tc_pen,
                            tc_tgt,
                            build_aum,
                            config,
                            weighting,
                        )

                    spec_tables: list[tuple[int | str, dict]] = []
                    for aum_case in aums:
                        aum, leg_aum_override, proportional_tc_only = leg_aum_for(
                            aum_case
                        )
                        # Weighted-cell nets for this spec; standard-cell nets
                        # are reused.
                        wt_nets = _compute_cell_nets(
                            wt_gross,
                            aum,
                            tc_context,
                            leg_aum_override,
                            proportional_tc_only,
                        )
                        cell_nets = {**std_nets_by_aum[aum_case], **wt_nets}
                        table = _decomposition_dict(
                            {cell: cell_nets[cell]["ls"] for cell in CELLS}
                        )
                        spec_tables.append((aum_case, table))

                        rows_13 = _leg_breakdown_rows(
                            model, weight_spec, aum_case, leg_capital, cell_nets,
                        )
                        rows_14 = _leg_decomposition_rows(
                            model, weight_spec, aum_case, leg_capital, cell_nets,
                        )
                        model_leg_rows.extend(rows_13)
                        model_leg_decomp_rows.extend(rows_14)
                        logger.info(
                            (
                                "[%s/%s/%s/%s] %s net SR: "
                                "1A=%.3f 1B=%.3f 1C=%.3f "
                                "2A=%.3f 2B=%.3f 2C=%.3f"
                            ),
                            stock_universe,
                            model,
                            weighting,
                            weight_spec,
                            _aum_title(mod22, aum_case),
                            table["SR_net_1A"],
                            table["SR_net_1B"],
                            table["SR_net_1C"],
                            table["SR_net_2A"],
                            table["SR_net_2B"],
                            table["SR_net_2C"],
                        )

                    if spec_tables:
                        tables_by_spec[weight_spec] = spec_tables

                if not tables_by_spec:
                    logger.warning(
                        "No specs produced for %s/%s/%s; skipping workbook",
                        stock_universe,
                        model,
                        weighting,
                    )
                    continue

                output_path = (
                    model_out_dir
                    / (
                        f"table12_two_sided_{leg_token}{weighting}.xlsx"
                    )
                )
                _write_table12_workbook(
                    mod22,
                    model,
                    tables_by_spec,
                    output_path,
                    weighting,
                    leg_capital,
                    stock_universe,
                )
                written.append(output_path)
                logger.info("Wrote %s", output_path)

                if model_leg_rows:
                    table13_path = (
                        model_out_dir
                        / (
                            f"table13_legs_two_sided_{leg_token}{weighting}.xlsx"
                        )
                    )
                    _write_table13_legs_workbook(
                        mod22,
                        model,
                        model_leg_rows,
                        table13_path,
                        weighting,
                        leg_capital,
                        stock_universe,
                    )
                    written.append(table13_path)
                    logger.info("Wrote %s", table13_path)

                if model_leg_decomp_rows:
                    table14_path = (
                        model_out_dir
                        / (
                            f"table14_legdecomp_two_sided_{leg_token}"
                            f"{weighting}.xlsx"
                        )
                    )
                    _write_table14_legdecomp_workbook(
                        mod22,
                        model,
                        model_leg_decomp_rows,
                        table14_path,
                        weighting,
                        leg_capital,
                        stock_universe,
                    )
                    written.append(table14_path)
                    logger.info("Wrote %s", table14_path)

    print("\n".join(str(p) for p in written) if written else "No outputs written.")


if __name__ == "__main__":
    main()
