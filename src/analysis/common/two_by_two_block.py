"""Shared Table-12-style 2x2 block renderer.

Single source of truth for the ``Panel A (Sharpe) / Panel B (2x2 decomposition,
optional ``*`` significance stars) / Panel C (diagnostics)`` workbook block used by
both the eval_realignment capacity tables (scripts 44/45) and the formal
long-short tables (scripts 22/22b). The ``*=LW p<0.05`` star convention and the
3-column layout live here so the two tracks stay byte-identical in presentation.
"""

from __future__ import annotations

import numpy as np
import pandas as pd


def set_two_by_two_widths(ws) -> None:
    for col, width in {"A": 36, "B": 20, "C": 20}.items():
        ws.column_dimensions[col].width = width


def write_two_by_two_block(
    ws,
    table: dict[str, float],
    gate_diag: dict[str, float],
    start_row: int,
    title: str,
    subtitle: str | None = None,
    column_headers: tuple[str, str] = ("Full rebalance", "Breakeven gate"),
    panel_c_title: str = "Panel C: Breakeven-Gate Diagnostics",
    panel_c_rows: list[tuple[str, float]] | None = None,
    star_effects: bool = False,
    include_panel_c: bool = True,
) -> int:
    """Table-12-style block: Panel A (Sharpe), B (2x2 decomposition), C (diagnostics).

    3-column layout: metric label plus the two cost-device columns. Defaults
    reproduce the script-44 breakeven block; callers with a different B device
    (e.g. the script-45 hysteresis book, or the formal long-short) override
    ``column_headers`` / ``panel_c_title`` / ``panel_c_rows``. Returns the next
    available row.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="404040")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    n_cols = 3

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = title_row + 1
    if subtitle:
        ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=n_cols)
        subtitle_cell = ws.cell(row=header_row, column=1, value=subtitle)
        subtitle_cell.font = Font(italic=True, size=10, color="555555")
        subtitle_cell.alignment = Alignment(horizontal="center")
        header_row += 1

    for col_idx, value in enumerate(["", *column_headers], start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=medium)
        cell.fill = header_fill

    panel_a_row = header_row + 1
    ws.merge_cells(start_row=panel_a_row, start_column=1, end_row=panel_a_row, end_column=n_cols)
    cell = ws.cell(row=panel_a_row, column=1, value="Panel A: Net Sharpe Ratios")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=thin)

    sr_rows = [
        ("Standard training (net)", ("1A", table["SR_net_1A"]), ("1B", table["SR_net_1B"])),
        ("Weighted training (net)", ("2A", table["SR_net_2A"]), ("2B", table["SR_net_2B"])),
        ("Standard training (gross)", ("1A", table["SR_gross_1A"]), ("1B", table["SR_gross_1B"])),
        ("Weighted training (gross)", ("2A", table["SR_gross_2A"]), ("2B", table["SR_gross_2B"])),
    ]
    row_idx = panel_a_row + 1
    for label, col_b, col_c in sr_rows:
        ws.cell(row=row_idx, column=1, value=label)
        for col_idx, (cell_name, value) in enumerate([col_b, col_c], start=2):
            ws.cell(row=row_idx, column=col_idx, value=f"{cell_name}: {value:.3f}")
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center"
            )
        row_idx += 1

    panel_b_row = row_idx
    ws.merge_cells(start_row=panel_b_row, start_column=1, end_row=panel_b_row, end_column=n_cols)
    panel_b_title = "Panel B: 2x2 Decomposition" + (
        "  (* = LW p < 0.05)" if star_effects else ""
    )
    cell = ws.cell(row=panel_b_row, column=1, value=panel_b_title)
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=medium)

    if star_effects:
        # Effect point estimates carrying a significance star (LW bootstrap
        # p < 0.05); the separate numeric p-value rows are omitted.
        def _starred(value, pval, fmt="{:.3f}"):
            if value is None or pd.isna(value):
                return ""
            text = fmt.format(float(value))
            if pd.notna(pval) and float(pval) < 0.05:
                text += "*"
            return text

        decomp_rows = [
            ("Training effect (2A - 1A)",
             _starred(table["training_effect"], table.get("lw_training_pval")), None),
            ("Portfolio effect (1B - 1A)",
             _starred(table["portfolio_effect"], table.get("lw_portfolio_pval")), None),
            ("Total effect (2B - 1A)",
             _starred(table["total_effect"], table.get("lw_total_pval")), None),
            ("Interaction", _starred(table["interaction"], None), None),
            ("Training share (%)", _starred(table["training_share_pct"], None, "{:.1f}"), None),
        ]
    else:
        decomp_rows = [
            ("Training effect (2A - 1A)", table["training_effect"], "0.000"),
            ("Portfolio effect (1B - 1A)", table["portfolio_effect"], "0.000"),
            ("Total effect (2B - 1A)", table["total_effect"], "0.000"),
            ("Interaction", table["interaction"], "0.000"),
            ("Training share (%)", table["training_share_pct"], "0.0"),
            ("LW p-val (training, net)", table["lw_training_pval"], "0.000"),
            ("LW p-val (total, net)", table["lw_total_pval"], "0.000"),
            ("LW p-val (H3, net)", table.get("lw_h3_pval", table.get("lw_h1_pval")), "0.000"),
        ]
    row_idx = panel_b_row + 1
    for label, value, fmt in decomp_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        if fmt is not None:
            value_cell.number_format = fmt
        value_cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    if include_panel_c:
        panel_c_row = row_idx
        ws.merge_cells(start_row=panel_c_row, start_column=1, end_row=panel_c_row, end_column=n_cols)
        cell = ws.cell(row=panel_c_row, column=1, value=panel_c_title)
        cell.font = Font(bold=True, italic=True)
        cell.border = Border(top=medium)

        diag_rows = panel_c_rows if panel_c_rows is not None else [
            ("Mean pass fraction (standard)", gate_diag.get("standard_pass_frac", np.nan)),
            ("Mean gross pass fraction (standard)", gate_diag.get("standard_gross_pass_frac", np.nan)),
            ("Mean pass fraction (weighted)", gate_diag.get("weighted_pass_frac", np.nan)),
            ("Mean gross pass fraction (weighted)", gate_diag.get("weighted_gross_pass_frac", np.nan)),
            ("Median |alpha| (standard)", gate_diag.get("standard_median_abs_alpha", np.nan)),
            ("Median |alpha| (weighted)", gate_diag.get("weighted_median_abs_alpha", np.nan)),
            ("Median half-spread", gate_diag.get("standard_median_half_spread", np.nan)),
        ]
        row_idx = panel_c_row + 1
        for label, value in diag_rows:
            ws.cell(row=row_idx, column=1, value=label)
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            value_cell.number_format = "0.0000"
            value_cell.alignment = Alignment(horizontal="center")
            row_idx += 1

    bottom_row = row_idx
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=bottom_row, column=col_idx).border = Border(top=medium)
    return bottom_row + 2
