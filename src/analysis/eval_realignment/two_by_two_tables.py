"""Old-format 2x2 deliverables for the capacity track (presentation layer).

Mirrors the formalanalysis portfolio deliverables (21e's ``two_by_three_{AUM}.csv``
long format and per-cell timeseries workbook, plus 22's formatted Table-12 block)
for the eval_realignment capacity books, with the 2x3 cell grid reduced to a 2x2:

    1A = standard training x full rebalance (the section-3 book, script 42)
    1B = standard training x breakeven gate (the section-4 book, script 43)
    2A = weighted training x full rebalance
    2B = weighted training x breakeven gate

The old metric names are kept verbatim for drop-in parity. In particular
``"Net portfolio effect annualized"`` here means the BREAKEVEN-GATE effect
SR(1B) - SR(1A) (the cost-device column effect), the analogue of the old
TC-aware-sort column effect; the tc-target (C) column is out of scope.

Inputs are the monthly series already written by scripts 42 and 43 — no panel is
loaded and no book is rebuilt. All four cells are aligned on their common months
before any cross-cell statistic.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

from src.evaluation.statistics import compute_effect_decomposition, sharpe_ratio

logger = logging.getLogger(__name__)

CELLS = ("1A", "1B", "2A", "2B")
MONTHS_PER_YEAR = 12

# (input file stem, row_type) -> cell
CELL_SOURCES = {
    "1A": ("capacity_portfolio_monthly", "standard"),
    "2A": ("capacity_portfolio_monthly", "weighted"),
    "1B": ("capacity_breakeven_monthly", "standard"),
    "2B": ("capacity_breakeven_monthly", "weighted"),
}

TIMESERIES_COLS = [
    "yyyymm", "gross_return", "transaction_cost", "net_return",
    "turnover", "n_long", "n_short",
]


def load_cell_series(
    spec_dir: Path, aum_label_str: str
) -> dict[str, pd.DataFrame] | None:
    """Read the 42/43 monthly CSVs and align the four cells on common months.

    Returns None (caller skips with a warning) when an input file or a row_type
    is missing, or when the cells share no months.
    """
    frames: dict[str, pd.DataFrame] = {}
    for stem in {src[0] for src in CELL_SOURCES.values()}:
        path = spec_dir / f"{stem}_{aum_label_str}.csv"
        if not path.exists():
            logger.warning("Missing input %s", path)
            return None
        frames[stem] = pd.read_csv(path)

    cells: dict[str, pd.DataFrame] = {}
    for cell, (stem, row_type) in CELL_SOURCES.items():
        sub = frames[stem][frames[stem]["row_type"] == row_type]
        if sub.empty:
            logger.warning(
                "No %s rows in %s_%s.csv", row_type, stem, aum_label_str
            )
            return None
        cells[cell] = sub.set_index("yyyymm").sort_index()

    common = sorted(set.intersection(*(set(df.index) for df in cells.values())))
    if not common:
        logger.warning("No common months across the four cells")
        return None
    return {cell: df.loc[common] for cell, df in cells.items()}


def _cell_summary(df: pd.DataFrame) -> dict[str, float]:
    """Per-cell return/TC/turnover summary on the aligned months (42 conventions)."""
    g = df["ret_gross"].to_numpy(dtype=np.float64)
    n = df["ret_net"].to_numpy(dtype=np.float64)
    tc = df["transaction_cost"]
    return {
        "gross_return_monthly": float(np.nanmean(g)),
        "gross_return_annual": float(np.nanmean(g)) * MONTHS_PER_YEAR,
        "net_return_monthly": float(np.nanmean(n)),
        "net_return_annual": float(np.nanmean(n)) * MONTHS_PER_YEAR,
        "tc_mean_monthly": float(tc.iloc[1:].mean()) if len(df) > 1 else np.nan,
        "tc_median_monthly": float(tc.iloc[1:].median()) if len(df) > 1 else np.nan,
        "gross_sr_monthly": sharpe_ratio(g),
        "net_sr_monthly": sharpe_ratio(n),
        "gross_sr_annual": sharpe_ratio(g, annualize=True),
        "net_sr_annual": sharpe_ratio(n, annualize=True),
        "turnover_mean": (
            float(df["turnover"].iloc[1:].mean()) if len(df) > 1 else np.nan
        ),
        "raw_trade_sum_mean": (
            float((2.0 * df["turnover"]).iloc[1:].mean()) if len(df) > 1 else np.nan
        ),
    }


def _safe_decomposition(
    cells: dict[str, pd.DataFrame],
    column: str,
    config: dict,
    yyyymm: np.ndarray | None,
) -> dict[str, Any]:
    """compute_effect_decomposition with a NaN fallback on short/degenerate samples."""
    try:
        return compute_effect_decomposition(
            returns_1a=cells["1A"][column].to_numpy(dtype=np.float64),
            returns_1b=cells["1B"][column].to_numpy(dtype=np.float64),
            returns_2a=cells["2A"][column].to_numpy(dtype=np.float64),
            returns_2b=cells["2B"][column].to_numpy(dtype=np.float64),
            yyyymm=yyyymm,
            config=config,
        )
    except (ValueError, KeyError, FileNotFoundError) as exc:
        logger.warning("Effect decomposition unavailable (%s): %s", column, exc)
        empty = {"p_value": np.nan}
        return {
            "lw_portfolio": empty, "lw_training": empty,
            "lw_total": empty, "lw_h3": empty, "factor_alphas": {},
        }


def format_two_by_two_rows(
    cells: dict[str, pd.DataFrame],
    config: dict,
    factor_alphas: bool = True,
) -> list[dict]:
    """Flatten the 2x2 result into metric/value rows (old two_by_three naming)."""
    summary = {cell: _cell_summary(cells[cell]) for cell in CELLS}
    net_sr = {c: summary[c]["net_sr_annual"] for c in CELLS}
    gross_sr = {c: summary[c]["gross_sr_annual"] for c in CELLS}

    net_training = net_sr["2A"] - net_sr["1A"]
    net_portfolio = net_sr["1B"] - net_sr["1A"]
    net_total = net_sr["2B"] - net_sr["1A"]
    net_interaction = net_total - net_training - net_portfolio
    gross_training = gross_sr["2A"] - gross_sr["1A"]
    gross_portfolio = gross_sr["1B"] - gross_sr["1A"]
    gross_total = gross_sr["2B"] - gross_sr["1A"]
    training_share = (
        net_training / net_total * 100.0 if abs(net_total) > 1e-10 else np.nan
    )

    yyyymm = cells["1A"].index.to_numpy() if factor_alphas else None
    decomp_net = _safe_decomposition(cells, "ret_net", config, yyyymm)
    decomp_gross = _safe_decomposition(cells, "ret_gross", config, None)

    rows: list[dict] = []
    for cell in CELLS:
        s = summary[cell]
        rows.extend(
            [
                {"metric": f"Gross return monthly ({cell})", "value": s["gross_return_monthly"]},
                {"metric": f"Gross return annualized ({cell})", "value": s["gross_return_annual"]},
                {"metric": f"Net return monthly ({cell})", "value": s["net_return_monthly"]},
                {"metric": f"Net return annualized ({cell})", "value": s["net_return_annual"]},
                {"metric": f"TC mean monthly ({cell})", "value": s["tc_mean_monthly"]},
                {"metric": f"TC median monthly ({cell})", "value": s["tc_median_monthly"]},
                {"metric": f"SR_gross_monthly({cell})", "value": s["gross_sr_monthly"]},
                {"metric": f"SR_net_monthly({cell})", "value": s["net_sr_monthly"]},
                {"metric": f"SR_gross_annualized({cell})", "value": s["gross_sr_annual"]},
                {"metric": f"SR_net_annualized({cell})", "value": s["net_sr_annual"]},
            ]
        )

    rows.extend(
        [
            {"metric": "Net training effect annualized", "value": net_training},
            {"metric": "Net portfolio effect annualized", "value": net_portfolio},
            {"metric": "Net total effect annualized", "value": net_total},
            {"metric": "Net interaction annualized", "value": net_interaction},
            {"metric": "Training share annualized (%)", "value": training_share},
            {"metric": "LW p-val (training, net)", "value": decomp_net["lw_training"].get("p_value", np.nan)},
            {"metric": "LW p-val (total, net)", "value": decomp_net["lw_total"].get("p_value", np.nan)},
            {"metric": "LW p-val (H1, net)", "value": decomp_net["lw_h3"].get("p_value", np.nan)},
            {"metric": "Gross training effect annualized", "value": gross_training},
            {"metric": "Gross portfolio effect annualized", "value": gross_portfolio},
            {"metric": "Gross total effect annualized", "value": gross_total},
            {"metric": "LW p-val (training, gross)", "value": decomp_gross["lw_training"].get("p_value", np.nan)},
            {"metric": "LW p-val (total, gross)", "value": decomp_gross["lw_total"].get("p_value", np.nan)},
            {"metric": "LW p-val (H1, gross)", "value": decomp_gross["lw_h3"].get("p_value", np.nan)},
        ]
    )

    alphas = decomp_net.get("factor_alphas", {})
    for model_name in ["capm", "ff3", "ff5", "ff5_mom"]:
        if model_name not in alphas:
            continue
        for cell in CELLS:
            alpha = alphas[model_name].get(cell, {})
            rows.extend(
                [
                    {"metric": f"alpha_{model_name}({cell})_annual", "value": alpha.get("alpha_annual", np.nan)},
                    {"metric": f"alpha_{model_name}({cell})_tstat", "value": alpha.get("alpha_tstat", np.nan)},
                    {"metric": f"alpha_{model_name}({cell})_pvalue", "value": alpha.get("alpha_pvalue", np.nan)},
                ]
            )

    for cell in CELLS:
        rows.extend(
            [
                {"metric": f"Turnover ({cell})", "value": summary[cell]["turnover_mean"]},
                {"metric": f"Raw trade sum ({cell})", "value": summary[cell]["raw_trade_sum_mean"]},
            ]
        )

    return rows


def write_timeseries_workbook(cells: dict[str, pd.DataFrame], path: Path) -> None:
    """One sheet per cell (1A/1B/2A/2B), old sheet column names, book-level only."""
    rename = {"ret_gross": "gross_return", "ret_net": "net_return"}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for cell in CELLS:
            sheet = cells[cell].reset_index().rename(columns=rename)
            sheet = sheet[[c for c in TIMESERIES_COLS if c in sheet.columns]]
            sheet.to_excel(writer, sheet_name=cell, index=False)
    logger.info("Saved %s", path)


def build_table_values(
    cells: dict[str, pd.DataFrame], config: dict
) -> dict[str, float]:
    """The value dict consumed by write_two_by_two_block (Table-12 analogue)."""
    summary = {cell: _cell_summary(cells[cell]) for cell in CELLS}
    net_sr = {c: summary[c]["net_sr_annual"] for c in CELLS}
    gross_sr = {c: summary[c]["gross_sr_annual"] for c in CELLS}
    training = net_sr["2A"] - net_sr["1A"]
    portfolio = net_sr["1B"] - net_sr["1A"]
    total = net_sr["2B"] - net_sr["1A"]
    decomp_net = _safe_decomposition(cells, "ret_net", config, None)
    return {
        **{f"SR_net_{c}": net_sr[c] for c in CELLS},
        **{f"SR_gross_{c}": gross_sr[c] for c in CELLS},
        "training_effect": training,
        "portfolio_effect": portfolio,
        "total_effect": total,
        "interaction": total - training - portfolio,
        "training_share_pct": (
            training / total * 100.0 if abs(total) > 1e-10 else np.nan
        ),
        "lw_training_pval": decomp_net["lw_training"].get("p_value", np.nan),
        "lw_total_pval": decomp_net["lw_total"].get("p_value", np.nan),
        "lw_h1_pval": decomp_net["lw_h3"].get("p_value", np.nan),
    }


def summarize_gate_diag(diag: pd.DataFrame) -> dict[str, float]:
    """Panel C values: mean pass fractions and scale medians per training row."""
    out: dict[str, float] = {}
    for row_type in ("standard", "weighted"):
        sub = diag[diag["row_type"] == row_type]
        prefix = row_type
        out[f"{prefix}_pass_frac"] = float(sub["pass_frac"].mean()) if len(sub) else np.nan
        out[f"{prefix}_gross_pass_frac"] = (
            float(sub["gross_pass_frac"].mean()) if len(sub) else np.nan
        )
        out[f"{prefix}_median_abs_alpha"] = (
            float(sub["median_abs_alpha"].median()) if len(sub) else np.nan
        )
        out[f"{prefix}_median_half_spread"] = (
            float(sub["median_half_spread"].median()) if len(sub) else np.nan
        )
    return out


def set_two_by_two_widths(ws) -> None:
    for col, width in {"A": 36, "B": 20, "C": 20}.items():
        ws.column_dimensions[col].width = width


def write_two_by_two_block(
    ws,
    table: dict[str, float],
    gate_diag: dict[str, float],
    start_row: int,
    title: str,
    subtitle: str | None = None,
) -> int:
    """Table-12-style block: Panel A (Sharpe), B (2x2 decomposition), C (gate diag).

    Mirrors the 22 formatter with a 3-column layout: metric label, full-rebalance
    column, breakeven-gate column. Returns the next available row.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    thin = Side(style="thin", color="808080")
    medium = Side(style="medium", color="404040")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    n_cols = 3

    title_row = start_row
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=n_cols)
    title_cell = ws.cell(row=title_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center")

    header_row = title_row + 1
    if subtitle:
        ws.merge_cells(start_row=header_row, start_column=1, end_row=header_row, end_column=n_cols)
        subtitle_cell = ws.cell(row=header_row, column=1, value=subtitle)
        subtitle_cell.font = Font(italic=True, size=10, color="555555")
        subtitle_cell.alignment = Alignment(horizontal="center")
        header_row += 1

    for col_idx, value in enumerate(["", "Full rebalance", "Breakeven gate"], start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=value)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(bottom=medium)
        cell.fill = header_fill

    panel_a_row = header_row + 1
    ws.merge_cells(start_row=panel_a_row, start_column=1, end_row=panel_a_row, end_column=n_cols)
    cell = ws.cell(row=panel_a_row, column=1, value="Panel A: Net Sharpe Ratios")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=thin)

    sr_rows = [
        ("Standard training (net)", ("1A", table["SR_net_1A"]), ("1B", table["SR_net_1B"])),
        ("Weighted training (net)", ("2A", table["SR_net_2A"]), ("2B", table["SR_net_2B"])),
        ("Standard training (gross)", ("1A", table["SR_gross_1A"]), ("1B", table["SR_gross_1B"])),
        ("Weighted training (gross)", ("2A", table["SR_gross_2A"]), ("2B", table["SR_gross_2B"])),
    ]
    row_idx = panel_a_row + 1
    for label, col_b, col_c in sr_rows:
        ws.cell(row=row_idx, column=1, value=label)
        for col_idx, (cell_name, value) in enumerate([col_b, col_c], start=2):
            ws.cell(row=row_idx, column=col_idx, value=f"{cell_name}: {value:.3f}")
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center"
            )
        row_idx += 1

    panel_b_row = row_idx
    ws.merge_cells(start_row=panel_b_row, start_column=1, end_row=panel_b_row, end_column=n_cols)
    cell = ws.cell(row=panel_b_row, column=1, value="Panel B: 2x2 Decomposition")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=medium)

    decomp_rows = [
        ("Training effect (2A - 1A)", table["training_effect"], "0.000"),
        ("Portfolio effect (1B - 1A)", table["portfolio_effect"], "0.000"),
        ("Total effect (2B - 1A)", table["total_effect"], "0.000"),
        ("Interaction", table["interaction"], "0.000"),
        ("Training share (%)", table["training_share_pct"], "0.0"),
        ("LW p-val (training, net)", table["lw_training_pval"], "0.000"),
        ("LW p-val (total, net)", table["lw_total_pval"], "0.000"),
        ("LW p-val (H1, net)", table["lw_h1_pval"], "0.000"),
    ]
    row_idx = panel_b_row + 1
    for label, value, fmt in decomp_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.number_format = fmt
        value_cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    panel_c_row = row_idx
    ws.merge_cells(start_row=panel_c_row, start_column=1, end_row=panel_c_row, end_column=n_cols)
    cell = ws.cell(row=panel_c_row, column=1, value="Panel C: Breakeven-Gate Diagnostics")
    cell.font = Font(bold=True, italic=True)
    cell.border = Border(top=medium)

    diag_rows = [
        ("Mean pass fraction (standard)", gate_diag.get("standard_pass_frac", np.nan)),
        ("Mean gross pass fraction (standard)", gate_diag.get("standard_gross_pass_frac", np.nan)),
        ("Mean pass fraction (weighted)", gate_diag.get("weighted_pass_frac", np.nan)),
        ("Mean gross pass fraction (weighted)", gate_diag.get("weighted_gross_pass_frac", np.nan)),
        ("Median |alpha| (standard)", gate_diag.get("standard_median_abs_alpha", np.nan)),
        ("Median |alpha| (weighted)", gate_diag.get("weighted_median_abs_alpha", np.nan)),
        ("Median half-spread", gate_diag.get("standard_median_half_spread", np.nan)),
    ]
    row_idx = panel_c_row + 1
    for label, value in diag_rows:
        ws.cell(row=row_idx, column=1, value=label)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=n_cols)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.number_format = "0.0000"
        value_cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    bottom_row = row_idx
    for col_idx in range(1, n_cols + 1):
        ws.cell(row=bottom_row, column=col_idx).border = Border(top=medium)
    return bottom_row + 2
